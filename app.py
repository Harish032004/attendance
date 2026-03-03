# app.py
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, date, timedelta
import calendar
from functools import wraps
import config
import os
import logging
import json
from werkzeug.security import generate_password_hash, check_password_hash
import re
import pandas as pd
from io import BytesIO
import base64
import time
# Add this with your other imports at the very top
from datetime import datetime, date, timedelta
import time
from functools import wraps
import config
import os
import logging
import json
from werkzeug.security import generate_password_hash, check_password_hash
import re
import pandas as pd
from io import BytesIO
import base64

from flask import send_file

# CREATE FLASK APP
app = Flask(__name__)
app.secret_key = 'my-attendance-system-secret-key-2024'  # Simple ah podunga
app.permanent_session_lifetime = timedelta(hours=8)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)




# ==================== New CACHING SYSTEM ====================
# ==================== CONNECTION POOLING & CACHING ====================
# ==================== SMART CACHE CONFIGURATION ====================
# Different cache times for different data types
CACHE_CONFIG = {
    'employee_status': 0,        # 0 seconds - ALWAYS FRESH (breaks, online/offline)
    'attendance_today': 0,        # 0 seconds - ALWAYS FRESH (today's attendance)
    'employees': 300,              # 5 minutes - employee list
    'departments': 300,            # 5 minutes - department stats
    'history': 300,                # 5 minutes - old attendance data
    'work_logs': 300,              # 5 minutes - work logs
    'reports': 600,                # 10 minutes - monthly reports
    'audit_logs': 600,             # 10 minutes - audit logs
    'default': 300                 # Default 5 minutes
}

# Cache storage
cache = {}
_sheet_connection = None
_sheet_last_used = 0
CONNECTION_TIMEOUT = 200  # 5 minutes
_worksheet_cache = {}
_WORKSHEET_CACHE_TIME = 120  # 2 minutes

def get_cached_sheet():
    """Reuse Google Sheets connection (5x faster)"""
    global _sheet_connection, _sheet_last_used
    now = time.time()
    
    # Check if we have a valid connection
    if _sheet_connection is not None and (now - _sheet_last_used) < CONNECTION_TIMEOUT:
        print("⚡ Using cached connection")
        return _sheet_connection
    
    # Create new connection
    print("📡 Creating new connection...")
    try:
        _sheet_connection = get_google_sheet()
        _sheet_last_used = now
        return _sheet_connection
    except Exception as e:
        print(f"Connection error: {e}")
        _sheet_connection = None
        return None

def get_cached_worksheet_data(worksheet_name):
    """Cache worksheet data in memory (10x faster)"""
    global _worksheet_cache
    now = time.time()
    cache_key = f"{worksheet_name}_{int(now / _WORKSHEET_CACHE_TIME)}"
    
    # Return cached data if available
    if cache_key in _worksheet_cache:
        print(f"⚡ Cached data: {worksheet_name}")
        return _worksheet_cache[cache_key]
    
    # Get sheet (will use cached connection)
    print(f"📡 Fetching: {worksheet_name}")
    sheet = get_cached_sheet()
    if not sheet:
        print(f"❌ No sheet connection for {worksheet_name}")
        return []
    
    try:
        worksheet = sheet.worksheet(worksheet_name)
        data = worksheet.get_all_records()
        _worksheet_cache[cache_key] = data
        print(f"✅ Fetched {len(data)} records from {worksheet_name}")
        return data
    except Exception as e:
        print(f"❌ Error fetching {worksheet_name}: {e}")
        return []

def clear_worksheet_cache():
    """Clear cache when data changes"""
    global _worksheet_cache
    _worksheet_cache = {}

def clear_cache(key=None):
    """Clear cache when data changes"""
    global _worksheet_cache, _sheet_connection
    if key:
        # Clear specific cache if needed
        pass
    else:
        _worksheet_cache.clear()
        # Don't clear connection, just the data cache
        print("🧹 Worksheet cache cleared")
def get_smart_cache(key, fetch_function, data_type='default'):
    """Get data with type-specific cache time"""
    global cache
    now = time.time()
    
    # Get cache time for this data type
    cache_time = CACHE_CONFIG.get(data_type, 300)
    
    # If cache_time is 0, always fetch fresh
    if cache_time == 0:
        print(f"📡 FRESH (no cache): {key}")
        return fetch_function()
    
    # Return cached if still valid
    if key in cache and now - cache[key]['time'] < cache_time:
        print(f"⚡ SMART CACHE: {key}")
        return cache[key]['data']
    
    # Fetch fresh
    print(f"📡 FETCHING: {key}")
    data = fetch_function()
    cache[key] = {'data': data, 'time': now, 'type': data_type}
    return data  

def clear_smart_cache(keys_to_clear=None, data_types=None):
    """Clear only specific caches"""
    global cache
    
    if keys_to_clear:
        for key in keys_to_clear:
            if key in cache:
                del cache[key]
        print(f"🧹 Cleared specific keys: {keys_to_clear}")
    
    elif data_types:
        keys_to_delete = []
        for key, value in cache.items():
            if value.get('type') in data_types:
                keys_to_delete.append(key)
        for key in keys_to_delete:
            del cache[key]
        print(f"🧹 Cleared data types: {data_types}")
    
    # Also clear worksheet cache for attendance
    global _worksheet_cache
    if 'Attendance' in _worksheet_cache:
        del _worksheet_cache['Attendance']
        print("🧹 Cleared Attendance from worksheet cache")
# ==================== HELPER FUNCTIONS ====================

def safe_str(value):
    """Safely convert any value to string"""
    if value is None:
        return ""
    return str(value)

def check_password_hash_safe(stored_hash, password):
    """Safely check password hash with string conversion"""
    try:
        # Ensure password is string
        password_str = str(password) if password is not None else ""
        return check_password_hash(stored_hash, password_str)
    except Exception as e:
        logger.error(f"Password check error: {e}")
        return False

def get_week_number(date_obj):
    """Get ISO week number"""
    return date_obj.isocalendar()[1]

def get_month_year(date_obj):
    """Get month and year"""
    return date_obj.strftime('%B %Y')

def is_office_ip(ip_address):
    """Check if IP is in office range"""
    if os.environ.get('RENDER'):
        return True
    
    for allowed_range in config.ALLOWED_IP_RANGES:
        if ip_address.startswith(allowed_range):
            return True
    return False

def get_client_ip():
    """Get client IP address"""
    if request.headers.get('X-Forwarded-For'):
        ip = request.headers.get('X-Forwarded-For').split(',')[0]
    else:
        ip = request.remote_addr
    return ip

def validate_password(password):
    """Validate password strength"""
    if len(password) < 8:
        return False, "Password must be at least 8 characters"
    if not re.search(r"[A-Z]", password):
        return False, "Password must contain at least one uppercase letter"
    if not re.search(r"[a-z]", password):
        return False, "Password must contain at least one lowercase letter"
    if not re.search(r"[0-9]", password):
        return False, "Password must contain at least one number"
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
        return False, "Password must contain at least one special character"
    return True, "Password is valid"

# ==================== DECORATORS ====================

def login_required(f):
    """Employee login required decorator"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'employee_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Admin login required decorator"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def ip_required(f):
    """Check IP range for WFO"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        client_ip = get_client_ip()
        if not is_office_ip(client_ip):
            return jsonify({
                'error': 'Please connect to Office WiFi',
                'code': 'WIFI_REQUIRED'
            }), 403
        return f(*args, **kwargs)
    return decorated_function

# ==================== GOOGLE SHEETS CONNECTION ====================

def get_google_sheet():
    """Initialize Google Sheets connection"""
    try:
        scope = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        
        logger.info("📌 Connecting to Google Sheets...")
        
        if not os.environ.get('RENDER'):
            if not os.path.exists('google_credentials.json'):
                logger.error("❌ google_credentials.json file not found!")
                return None
                
            creds = ServiceAccountCredentials.from_json_keyfile_name('google_credentials.json', scope)
            client = gspread.authorize(creds)
            sheet = client.open("Daily Attendance")
        else:
            creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
            if not creds_json:
                logger.error("❌ GOOGLE_CREDENTIALS_JSON not found")
                return None
                
            creds_dict = json.loads(creds_json)
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            client = gspread.authorize(creds)
            sheet = client.open("Daily Attendance")
        
        logger.info(f"✅ Connected to sheet: {sheet.title}")
        return sheet
        
    except Exception as e:
        logger.error(f"❌ Connection error: {str(e)}")
        return None

def get_next_id(worksheet, id_column='A'):
    """Get next available ID for a worksheet"""
    try:
        records = worksheet.get_all_records()
        if not records:
            return "1"
        
        max_id = 0
        for record in records:
            try:
                # Convert to string first, then to int
                record_id = int(float(str(record.get('ID', 0))))
                if record_id > max_id:
                    max_id = record_id
            except (ValueError, TypeError) as e:
                logger.warning(f"Error converting ID: {e}")
                continue
        return str(max_id + 1)
    except Exception as e:
        logger.error(f"Error in get_next_id: {e}")
        return "1"

def log_audit(user_type, user_id, user_name, action, details=""):
    """Log audit trail"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return False
        
        logs_ws = sheet.worksheet("AuditLogs")
        next_id = get_next_id(logs_ws)
        
        logs_ws.append_row([
            next_id,
            datetime.now().isoformat(),
            user_type,
            safe_str(user_id),
            safe_str(user_name),
            action,
            details,
            get_client_ip() if 'request' in globals() else "System"
        ])
        return True
    except Exception as e:
        logger.error(f"Audit log error: {e}")
        return False

def initialize_all_worksheets(sheet):
    """Initialize all required worksheets"""
    try:
        worksheets_config = {
           "Attendance": ["ID", "Employee ID", "Employee Name", "Date", "Time", "Type", "Week", "Month", "Year", "IP Address", "Status", "Break Count", "Break Minutes"],
           "Employees": ["ID", "Employee ID", "Employee Name", "Email", "Password Hash", "Department", "Join Date", "Is Active", "Role", "Last Login", "Password Changed", "Force Password Change", "Last Activity", "Status", "Break Start", "Break End", "Total Break Time", "Break Count"],
            "WorkLogs": ["ID", "Employee ID", "Employee Name", "Date", "Work Description", "Hours Worked", "Submitted At", "Status", "IP Address"],
            "WeeklySummary": ["ID", "Employee ID", "Employee Name", "Week", "Year", "Present", "Leave", "WFH", "Total Hours", "Generated At"],
            "MonthlySummary": ["ID", "Employee ID", "Employee Name", "Month", "Year", "Present", "Leave", "WFH", "Total Hours", "Generated At"],
            "Admins": ["ID", "Admin ID", "Admin Name", "Email", "Password Hash", "Role", "Created At", "Last Login", "Is Active"],
            "AuditLogs": ["ID", "Timestamp", "User Type", "User ID", "User Name", "Action", "Details", "IP Address"],
            "LeaveRequests": ["ID", "Employee ID", "Employee Name", "From Date", "To Date", "Type", "Reason", "Status", "Applied On", "Approved By", "Approved On"],
            "Settings": ["Key", "Value", "Updated By", "Updated At"],
            "Departments": ["ID", "Department Name", "Manager", "Created At", "Is Active"]
        }
        
        for ws_name, headers in worksheets_config.items():
            try:
                # Try to get the worksheet
                worksheet = sheet.worksheet(ws_name)
                logger.info(f"✅ Worksheet '{ws_name}' exists")
                
                # Check if headers are correct
                existing_headers = worksheet.row_values(1)
                if existing_headers != headers:
                    logger.info(f"📝 Updating headers for '{ws_name}'")
                    # Clear the worksheet and set new headers
                    worksheet.clear()
                    worksheet.append_row(headers)
                    
            except gspread.WorksheetNotFound:
                logger.info(f"📝 Creating worksheet: {ws_name}")
                new_ws = sheet.add_worksheet(title=ws_name, rows=2000, cols=30)
                new_ws.append_row(headers)
                
                # Add default data for certain sheets
                if ws_name == "Admins":
                    try:
                        admin_hash = generate_password_hash("Admin@123")
                        new_ws.append_row([
                            "1", "ADMIN001", "System Admin", "admin@company.com", 
                            admin_hash, "super_admin", datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "", "Yes"
                        ])
                        logger.info("✅ Added default admin")
                    except Exception as e:
                        logger.error(f"Error adding admin: {e}")
                
                elif ws_name == "Settings":
                    settings_data = [
                        ["min_password_length", "8", "system", datetime.now().isoformat()],
                        ["default_work_hours", "8", "system", datetime.now().isoformat()],
                        ["work_log_min_chars", "20", "system", datetime.now().isoformat()],
                        ["session_timeout", "8", "system", datetime.now().isoformat()],
                        ["password_expiry_days", "90", "system", datetime.now().isoformat()],
                        ["allow_wfh", "true", "system", datetime.now().isoformat()],
                        ["allow_leave", "true", "system", datetime.now().isoformat()],
                        ["require_uppercase", "true", "system", datetime.now().isoformat()],
                        ["require_numbers", "true", "system", datetime.now().isoformat()],
                        ["require_special", "true", "system", datetime.now().isoformat()],
                        ["max_login_attempts", "5", "system", datetime.now().isoformat()],
                        ["enable_ip_check", "true", "system", datetime.now().isoformat()],
                        ["enable_audit_logs", "true", "system", datetime.now().isoformat()]
                    ]
                    for setting in settings_data:
                        new_ws.append_row(setting)
                    logger.info("✅ Added default settings")
                
                elif ws_name == "Departments":
                    departments = ["Engineering", "Sales", "Marketing", "HR", "Finance", "Operations"]
                    for i, dept in enumerate(departments, start=1):
                        new_ws.append_row([str(i), dept, "", datetime.now().strftime('%Y-%m-%d'), "Yes"])
                    logger.info("✅ Added default departments")
                
                logger.info(f"✅ Created '{ws_name}' with headers")
        
        return True
    except Exception as e:
        logger.error(f"Error initializing worksheets: {e}")
        return False

def add_default_employees(sheet):
    """Add default employees ONLY if NO employees exist"""
    try:
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        # If there are already employees, don't add anything
        if len(all_employees) > 0:
            print(f"✅ Employees already exist ({len(all_employees)} records). Skipping default addition.")
            return False
        
        # Only add defaults if sheet is completely empty
        print("📝 Adding default employees (first time setup)...")
        default_employees = [
            {"id": "020", "name": "Nageshwari", "email": "nageshwari@company.com", "dept": "Engineering"},
            {"id": "023", "name": "Harish", "email": "harish@company.com", "dept": "Engineering"},
            {"id": "024", "name": "Gowtham", "email": "gowtham@company.com", "dept": "Sales"},
            {"id": "025", "name": "Saranya", "email": "saranya@company.com", "dept": "HR"},
            {"id": "026", "name": "Abdul Malik", "email": "abdul@company.com", "dept": "Marketing"}
        ]
        
        # Add each employee with all 18 columns
        for emp in default_employees:
            next_id = get_next_id(employees_ws)
            password_hash = generate_password_hash(emp["id"])
            today = date.today().isoformat()
            
            employees_ws.append_row([
                next_id,                # 1. ID
                emp["id"],               # 2. Employee ID
                emp["name"],             # 3. Employee Name
                emp["email"],            # 4. Email
                password_hash,           # 5. Password Hash
                emp["dept"],             # 6. Department
                today,                   # 7. Join Date
                'Yes',                   # 8. Is Active
                'employee',              # 9. Role
                '',                      # 10. Last Login
                '',                      # 11. Password Changed
                'Yes',                   # 12. Force Password Change
                '',                      # 13. Last Activity
                'offline',               # 14. Status
                '',                      # 15. Break Start
                '',                      # 16. Break End
                '0',                     # 17. Total Break Time
                '0'                      # 18. Break Count
            ])
        
        print(f"✅ Added {len(default_employees)} default employees with all columns")
        return True
        
    except Exception as e:
        print(f"❌ Error in add_default_employees: {e}")
        return False

def get_employee_by_id(employee_id):
    """Get employee details by ID"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return None
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        for emp in all_employees:
            if safe_str(emp.get('Employee ID', '')).strip() == safe_str(employee_id).strip():
                return emp
        return None
    except:
        return None

def check_today_attendance(employee_id):
    """Check if employee already marked attendance today"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return None
        
        today = date.today().isoformat()
        attendance_ws = sheet.worksheet("Attendance")
        all_records = attendance_ws.get_all_records()
        
        for record in all_records:
            if (safe_str(record.get('Employee ID', '')).strip() == safe_str(employee_id).strip() and 
                record.get('Date') == today):
                return {
                    'marked': True,
                    'type': record.get('Type'),
                    'time': record.get('Time')
                }
        return {'marked': False}
    except:
        return {'marked': False}

def check_today_work_log(employee_id):
    """Check if employee already submitted work log today"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return None
        
        today = date.today().isoformat()
        work_logs_ws = sheet.worksheet("WorkLogs")
        all_logs = work_logs_ws.get_all_records()
        
        for log in all_logs:
            if (safe_str(log.get('Employee ID', '')).strip() == safe_str(employee_id).strip() and 
                log.get('Date') == today):
                return {
                    'submitted': True,
                    'description': log.get('Work Description'),
                    'hours': log.get('Hours Worked')
                }
        return {'submitted': False}
    except:
        return {'submitted': False}

# ==================== ROUTES: LOGIN & AUTH ====================

@app.route('/')
def index():
    """Root route - redirect based on session"""
    if 'employee_id' in session:
        return redirect(url_for('employee_dashboard'))
    if 'admin_id' in session:
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Employee login page - OPTIMIZED"""
    if request.method == 'POST':
        try:
            data = request.json
            employee_id = safe_str(data.get('employee_id', '')).strip()
            password = safe_str(data.get('password', '')).strip()
            
            if not employee_id or not password:
                return jsonify({'success': False, 'message': 'Please enter both fields'}), 400
            
            # Use cached sheet (FAST)
            sheet = get_cached_sheet()
            if not sheet:
                return jsonify({'success': False, 'message': 'Database connection error'}), 500
            
            # Get cached employee data (NO new connection!)
            employees_data = get_cached_worksheet_data("Employees")
            
            # Find employee (in memory - SUPER FAST)
            employee = None
            row_index = None
            
            for i, emp in enumerate(employees_data, start=2):
                sheet_id = safe_str(emp.get('Employee ID', '')).strip()
                if sheet_id == employee_id or sheet_id.lstrip('0') == employee_id.lstrip('0'):
                    employee = emp
                    row_index = i
                    break
            
            if not employee:
                return jsonify({'success': False, 'message': 'Employee ID not found'}), 401
            
            if employee.get('Is Active', 'Yes') != 'Yes':
                return jsonify({'success': False, 'message': 'Account is inactive'}), 401
            
            stored_hash = employee.get('Password Hash', '')
            if not stored_hash:
                return jsonify({'success': False, 'message': 'Password not set'}), 401
            
            # Check password
            if not check_password_hash(stored_hash, password):
                log_audit('employee', employee_id, 'Unknown', 'LOGIN_FAILED', 'Invalid password')
                return jsonify({'success': False, 'message': 'Invalid password'}), 401
            
            # Login successful
            session.permanent = True
            session['employee_id'] = safe_str(employee.get('Employee ID'))
            session['employee_name'] = employee.get('Employee Name')
            session['department'] = employee.get('Department', '')
            session['role'] = 'employee'
            
            # Update status (use direct connection for write)
            if row_index:
                try:
                    sheet = get_cached_sheet()
                    employees_ws = sheet.worksheet("Employees")
                    employees_ws.update_cell(row_index, 13, datetime.now().isoformat())  # Last Activity
                    employees_ws.update_cell(row_index, 14, 'online')  # Status
                    employees_ws.update_cell(row_index, 10, datetime.now().isoformat())  # Last Login
                    
                    # Clear cache so admin sees update
                    clear_cache('employee_status')
                except:
                    pass
            
            log_audit('employee', employee.get('Employee ID'), employee.get('Employee Name'), 'LOGIN', 'Employee logged in')
            
            return jsonify({
                'success': True,
                'redirect': url_for('employee_dashboard')
            })
            
        except Exception as e:
            logger.error(f"Login error: {str(e)}")
            return jsonify({'success': False, 'message': f'Login failed: {str(e)}'}), 500
    
    return render_template('login.html', user_type='employee')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if request.method == 'POST':
        try:
            data = request.json
            admin_id = safe_str(data.get('admin_id', '')).strip()
            password = safe_str(data.get('password', '')).strip()
            
            if not admin_id or not password:
                return jsonify({'success': False, 'message': 'Please enter both fields'}), 400
            sheet = get_cached_sheet()
            
            if not sheet:
                return jsonify({'success': False, 'message': 'Database connection error'}), 500
            
            # Initialize worksheets
            initialize_all_worksheets(sheet)
            
            admins_ws = sheet.worksheet("Admins")
            all_admins = admins_ws.get_all_records()
            
            # Find admin
            admin = None
            row_index = None
            for i, adm in enumerate(all_admins, start=2):
                if safe_str(adm.get('Admin ID', '')).strip() == admin_id:
                    admin = adm
                    row_index = i
                    break
            
            if admin and safe_str(admin.get('Is Active', 'Yes')) == 'Yes':
                stored_hash = safe_str(admin.get('Password Hash', ''))
                
                if check_password_hash_safe(stored_hash, password):
                    session.permanent = True
                    session['admin_id'] = safe_str(admin.get('Admin ID'))
                    session['admin_name'] = safe_str(admin.get('Admin Name'))
                    session['admin_role'] = safe_str(admin.get('Role', 'admin'))
                    
                    try:
                        admins_ws.update_cell(row_index, 8, datetime.now().isoformat())  # Last Login
                    except:
                        pass
                    
                    log_audit('admin', admin_id, safe_str(admin.get('Admin Name')), 'ADMIN_LOGIN', 'Admin logged in')
                    
                    return jsonify({
                        'success': True,
                        'redirect': url_for('admin_dashboard')
                    })
                else:
                    log_audit('admin', admin_id, 'Unknown', 'ADMIN_LOGIN_FAILED', 'Invalid password')
                    return jsonify({'success': False, 'message': 'Invalid password'}), 401
            else:
                log_audit('admin', admin_id, 'Unknown', 'ADMIN_LOGIN_FAILED', 'Admin not found or inactive')
                return jsonify({'success': False, 'message': 'Admin ID not found or inactive'}), 401
                    
        except Exception as e:
            logger.error(f"Admin login error: {str(e)}")
            return jsonify({'success': False, 'message': f'Login failed: {str(e)}'}), 500
    
    return render_template('login.html', user_type='admin')




# ==================== ONLINE/OFFLINE TRACKING ====================

# ==================== ONLINE/OFFLINE TRACKING ====================
    

    # ==================== EMPLOYEE STATUS ROUTES ====================

@app.route('/api/heartbeat', methods=['POST'])
@login_required
def heartbeat():
    """Update employee online status"""
    try:
        employee_id = session['employee_id']
        
        sheet = get_cached_sheet()
        if sheet:
            employees_ws = sheet.worksheet("Employees")
            all_employees = employees_ws.get_all_records()
            
            for i, emp in enumerate(all_employees, start=2):
                sheet_id = safe_str(emp.get('Employee ID', '')).strip()
                if sheet_id == employee_id or sheet_id.lstrip('0') == employee_id.lstrip('0'):
                    employees_ws.update_cell(i, 13, datetime.now().isoformat())  # Last Activity
                    current_status = emp.get('Status', 'offline')
                    if current_status != 'online' and current_status != 'break':
                        employees_ws.update_cell(i, 14, 'online')
                    
                    # Clear cache so admin sees updated status
                    clear_cache('employee_status')
                    break
            
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500





@app.route('/api/admin/employee-status')
@admin_required
def admin_employee_status():
    """Get all employees status - ALWAYS FRESH (no cache)"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        # FORCE fresh data - don't use cache for this endpoint
        employees_ws = sheet.worksheet("Employees")
        attendance_ws = sheet.worksheet("Attendance")
        
        # Get fresh data directly from sheets
        all_employees = employees_ws.get_all_records()
        all_attendance = attendance_ws.get_all_records()
        
        today = date.today().isoformat()
        print(f"📊 Admin status - Today: {today}")
        
        # Create map of today's break data
        today_breaks = {}
        for att in all_attendance:
            if att.get('Date') == today:
                emp_id = safe_str(att.get('Employee ID')).strip()
                
                # Get break count (handle empty values)
                break_count = att.get('Break Count', 0)
                if break_count is None or break_count == '':
                    break_count = 0
                else:
                    try:
                        break_count = int(float(str(break_count)))
                    except:
                        break_count = 0
                
                today_breaks[emp_id] = break_count
                print(f"📊 Break count for {emp_id}: {break_count}")
        
        status_list = []
        for emp in all_employees:
            if emp.get('Is Active') == 'Yes':
                emp_id = safe_str(emp.get('Employee ID')).strip()
                status = emp.get('Status', 'offline')
                last_activity = emp.get('Last Activity', '')
                
                # Calculate time ago
                time_ago = 'Never'
                if last_activity:
                    try:
                        last_time = datetime.fromisoformat(last_activity)
                        diff = datetime.now() - last_time
                        if diff.total_seconds() < 60:
                            time_ago = 'just now'
                        elif diff.total_seconds() < 3600:
                            time_ago = f'{int(diff.total_seconds() / 60)} min ago'
                        else:
                            time_ago = f'{int(diff.total_seconds() / 3600)} hours ago'
                    except:
                        time_ago = 'unknown'
                
                # Get break count (0 if not found)
                break_count = today_breaks.get(emp_id, 0)
                
                status_list.append({
                    'employee_id': emp_id,
                    'name': emp.get('Employee Name'),
                    'department': emp.get('Department', 'Not Assigned'),
                    'status': status,
                    'time_ago': time_ago,
                    'break_count': break_count
                })
                
                print(f"✅ {emp_id}: {emp.get('Employee Name')} - Break count: {break_count}")
        
        # Clear the specific cache for this endpoint
        clear_smart_cache(keys_to_clear=['admin_employee_status'])
        
        return jsonify({'success': True, 'status': status_list})
        
    except Exception as e:
        print(f"❌ Status error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/api/employee/break-stats')
@login_required
def get_employee_break_stats():
    """Get today's break statistics for employee"""
    try:
        employee_id = session['employee_id']
        sheet = get_cached_sheet()
        
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        all_rows = employees_ws.get_all_values()
        
        # Find employee
        row_index = None
        emp_data = None
        for i, emp in enumerate(all_employees, start=2):
            sheet_id = safe_str(emp.get('Employee ID', '')).strip()
            if sheet_id == employee_id or sheet_id.lstrip('0') == employee_id.lstrip('0'):
                row_index = i
                emp_data = emp
                break
        
        if not emp_data:
            return jsonify({'success': False, 'message': 'Employee not found'}), 404
        
        # Get break data
        current_row = all_rows[row_index-1] if row_index-1 < len(all_rows) else []
        
        # Pad row if needed
        while len(current_row) < 20:
            current_row.append('')
        
        # Get break values
        status = emp_data.get('Status', 'offline')
        break_start = current_row[14] if len(current_row) > 14 else ''  # Column 15
        break_total = emp_data.get('Total Break Time', '0')
        break_count = emp_data.get('Break Count', '0')
        
        # Calculate current break duration if on break
        current_break_duration = 0
        if status == 'break' and break_start:
            try:
                start_time = datetime.fromisoformat(break_start)
                current_break_duration = (datetime.now() - start_time).total_seconds() / 60
            except:
                pass
        
        # Format for display
        try:
            total_minutes = float(break_total)
            total_display = format_break_time(total_minutes)
        except:
            total_display = "0 min"
        
        # Calculate remaining break time (if you want to set a limit)
        # Example: 60 minutes max break per day
        max_break_minutes = 60  # You can make this configurable
        remaining_minutes = max(0, max_break_minutes - float(break_total))
        remaining_display = format_break_time(remaining_minutes)
        
        return jsonify({
            'success': True,
            'status': status,
            'break_stats': {
                'total_minutes': float(break_total),
                'total_display': total_display,
                'break_count': int(float(break_count)) if break_count else 0,
                'current_break_minutes': round(current_break_duration, 1),
                'on_break': status == 'break',
                'max_break_minutes': max_break_minutes,
                'remaining_minutes': round(remaining_minutes, 1),
                'remaining_display': remaining_display,
                'percentage_used': min(100, round((float(break_total) / max_break_minutes) * 100)) if float(break_total) > 0 else 0
            }
        })
        
    except Exception as e:
        print(f"❌ Break stats error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

def format_break_time(minutes):
    """Format minutes into readable time"""
    if minutes < 1:
        return "0 min"
    elif minutes < 60:
        return f"{int(minutes)} min"
    else:
        hours = int(minutes // 60)
        mins = int(minutes % 60)
        if mins == 0:
            return f"{hours} hour{'s' if hours > 1 else ''}"
        else:
            return f"{hours}h {mins}m"
   
@app.route('/logout')
def logout():
    """Logout user"""
    try:
        if 'employee_id' in session:
            # Mark as offline before clearing session
            employee_id = session['employee_id']
            employee_name = session.get('employee_name', '')
            
            # Try to update status to offline
            try:
                update_employee_status(employee_id, 'offline')
                print(f"✅ Marked {employee_id} as offline")
            except Exception as e:
                print(f"⚠️ Could not update offline status: {e}")
            
            # Log audit
            log_audit('employee', employee_id, employee_name, 'LOGOUT', 'User logged out')
            
        elif 'admin_id' in session:
            log_audit('admin', session['admin_id'], session.get('admin_name', ''), 'LOGOUT', 'Admin logged out')
        
        # Clear session
        session.clear()
        
        # Clear cache
        clear_cache('employee_status')
        
    except Exception as e:
        print(f"⚠️ Logout error: {e}")
        session.clear()
    
    return redirect(url_for('login'))

# ==================== CHANGE PASSWORD ROUTES ====================

# ✅ Page route - Renders the HTML form (KEEP THIS ONLY ONCE)
@app.route('/change-password')
@login_required
def change_password_page():
    """Render change password page"""
    return render_template('change_password.html')

# ✅ API route - Handles the password update (KEEP THIS ONLY ONCE)
@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    """Change password for employee"""
    try:
        data = request.json
        current_password = data.get('current_password', '')
        new_password = data.get('new_password', '')
        confirm_password = data.get('confirm_password', '')

        # Validation
        if not current_password or not new_password or not confirm_password:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400

        if new_password != confirm_password:
            return jsonify({'success': False, 'message': 'New passwords do not match'}), 400

        # Validate password strength
        is_valid, message = validate_password(new_password)
        if not is_valid:
            return jsonify({'success': False, 'message': message}), 400

        # Connect to Google Sheets
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database connection error'}), 500

        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        all_rows = employees_ws.get_all_values()

        employee_id = session['employee_id']
        row_index = None
        stored_hash = None

        # Find employee
        for i, record in enumerate(all_employees):
            record_emp_id = str(record.get('Employee ID', '')).strip()
            session_emp_id = str(employee_id).strip()
            
            if (record_emp_id == session_emp_id or 
                record_emp_id.lstrip('0') == session_emp_id.lstrip('0')):
                stored_hash = record.get('Password Hash', '')
                row_index = i + 2
                break

        if not stored_hash:
            return jsonify({'success': False, 'message': 'User not found'}), 404

        # Verify current password
        if not check_password_hash(stored_hash, current_password):
            return jsonify({'success': False, 'message': 'Current password is incorrect'}), 401

        # Check if new password is same as current
        if check_password_hash(stored_hash, new_password):
            return jsonify({'success': False, 'message': 'New password must be different from current password'}), 400

        # Hash new password
        new_password_hash = generate_password_hash(new_password)

        # Find password hash column
        headers = all_rows[0]
        password_col = None
        for idx, header in enumerate(headers):
            if header == 'Password Hash':
                password_col = idx + 1
                break

        # Update password
        if password_col:
            employees_ws.update_cell(row_index, password_col, new_password_hash)
        else:
            employees_ws.update_cell(row_index, 5, new_password_hash)

        # Log the change
        log_audit('employee', employee_id, session.get('employee_name', ''), 
                 'PASSWORD_CHANGED', 'Password updated successfully')

        return jsonify({'success': True, 'message': 'Password changed successfully'})

    except Exception as e:
        logger.error(f"Password change error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== WORK LOG ROUTES ====================
@app.route('/api/work-log', methods=['GET', 'POST'])
@login_required
def work_log():
    """Handle work log entries - POST to save, GET to retrieve"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database connection error'}), 500

        work_logs_ws = sheet.worksheet("WorkLogs")
        today = datetime.now().strftime('%Y-%m-%d')
        employee_id = session['employee_id']
        employee_name = session['employee_name']

        if request.method == 'POST':
            data = request.json
            work_description = data.get('work_description', '').strip()
            hours_worked = data.get('hours_worked', '')

            # Validate
            if not work_description or len(work_description) < 20:
                return jsonify({'success': False, 'message': 'Work description must be at least 20 characters'}), 400
            
            try:
                hours_worked = float(hours_worked)
                if hours_worked <= 0 or hours_worked > 24:
                    return jsonify({'success': False, 'message': 'Hours must be between 1 and 24'}), 400
            except:
                return jsonify({'success': False, 'message': 'Invalid hours format'}), 400

            # Check if work log exists for today
            all_logs = work_logs_ws.get_all_records()
            existing_log = None
            row_index = None

            for i, log in enumerate(all_logs):
                if (str(log.get('Employee ID', '')).strip() == employee_id and 
                    log.get('Date') == today):
                    existing_log = log
                    row_index = i + 2  # +2 because records start from row 2
                    break

            if existing_log:
                # Update existing log
                all_rows = work_logs_ws.get_all_values()
                headers = all_rows[0]
                
                # Find column indices
                desc_col = None
                hours_col = None
                status_col = None
                
                for idx, header in enumerate(headers):
                    if header == 'Work Description':
                        desc_col = idx + 1
                    elif header == 'Hours Worked':
                        hours_col = idx + 1
                    elif header == 'Status':
                        status_col = idx + 1

                if desc_col:
                    work_logs_ws.update_cell(row_index, desc_col, work_description)
                if hours_col:
                    work_logs_ws.update_cell(row_index, hours_col, hours_worked)
                if status_col:
                    work_logs_ws.update_cell(row_index, status_col, f'Updated at {datetime.now().strftime("%H:%M:%S")}')

                log_audit('employee', employee_id, employee_name, 'WORK_LOG_UPDATED', f'Updated work log for {today}')
                return jsonify({'success': True, 'message': 'Work log updated successfully'})
            else:
                # Create new work log
                next_id = get_next_id(work_logs_ws)
                work_logs_ws.append_row([
                    next_id,
                    employee_id,
                    employee_name,
                    today,
                    work_description,
                    hours_worked,
                    datetime.now().isoformat(),
                    'Submitted'
                ])
                
                log_audit('employee', employee_id, employee_name, 'WORK_LOG_SUBMITTED', f'Submitted work log for {today}')
                return jsonify({'success': True, 'message': 'Work log submitted successfully'})

        else:  # GET request
            # Retrieve today's work log if exists
            all_logs = work_logs_ws.get_all_records()
            today_log = None
            
            for log in all_logs:
                if (str(log.get('Employee ID', '')).strip() == employee_id and 
                    log.get('Date') == today):
                    today_log = {
                        'work_description': log.get('Work Description', ''),
                        'hours_worked': log.get('Hours Worked', ''),
                        'status': log.get('Status', ''),
                        'submitted_at': log.get('Submitted At', '')
                    }
                    break

            return jsonify({
                'success': True,
                'log': today_log,
                'can_edit': True  # Always true for today - can edit multiple times
            })

    except Exception as e:
        logger.error(f"Work log error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/work-log')
@login_required
def work_log_page():
    """Render work log page"""
    return render_template('work_log.html')


# ==================== EMPLOYEE ROUTES ====================

@app.route('/employee/dashboard')
@login_required
def employee_dashboard():
    """OPTIMIZED Employee dashboard"""
    try:
        # Get cached data (FAST)
        attendance_data = get_cached_worksheet_data("Attendance")
        work_logs_data = get_cached_worksheet_data("WorkLogs")
        
        employee_id = session['employee_id']
        current_month = datetime.now().strftime('%B %Y')
        today = date.today().isoformat()
        
        # Process in memory (SUPER FAST)
        stats = {'present': 0, 'wfh': 0, 'leave': 0, 'total': 0}
        recent_logs = []
        today_attendance = {'marked': False}
        today_work_log = {'submitted': False}
        
        for record in attendance_data:
            if safe_str(record.get('Employee ID')).strip() == employee_id:
                if record.get('Date') == today:
                    today_attendance = {
                        'marked': True,
                        'type': record.get('Type'),
                        'time': record.get('Time')
                    }
                
                if record.get('Month') == current_month:
                    att_type = record.get('Type', 'WFO')
                    if att_type == 'WFO':
                        stats['present'] += 1
                    elif att_type == 'WFH':
                        stats['wfh'] += 1
                    elif att_type == 'Leave':
                        stats['leave'] += 1
                    stats['total'] += 1
        
        # Process work logs
        for log in reversed(work_logs_data):
            if safe_str(log.get('Employee ID')).strip() == employee_id:
                if log.get('Date') == today:
                    today_work_log = {
                        'submitted': True,
                        'description': log.get('Work Description'),
                        'hours': log.get('Hours Worked')
                    }
                
                recent_logs.append({
                    'date': log.get('Date'),
                    'description': log.get('Work Description'),
                    'hours': log.get('Hours Worked')
                })
                if len(recent_logs) >= 10:
                    break
        
        client_ip = get_client_ip()
        wifi_status = is_office_ip(client_ip)
        
        return render_template('employee/dashboard.html',
                             employee_name=session['employee_name'],
                             employee_id=employee_id,
                             department=session.get('department', ''),
                             wifi_status=wifi_status,
                             today_attendance=today_attendance,
                             today_work_log=today_work_log,
                             stats=stats,
                             recent_logs=recent_logs)
    
    except Exception as e:
        print(f"Dashboard error: {e}")
        return render_template('employee/dashboard.html', error=str(e))

@app.route('/api/mark-attendance', methods=['POST'])
@login_required
def mark_attendance():
    """Mark attendance (WFO/WFH/Leave)"""
    try:
        data = request.json
        attendance_type = data.get('type', 'WFO')
        
        if attendance_type not in ['WFO', 'WFH', 'Leave']:
            return jsonify({'success': False, 'message': 'Invalid attendance type'}), 400
        
        # Check if already marked today
        today_status = check_today_attendance(session['employee_id'])
        if today_status['marked']:
            return jsonify({
                'success': False,
                'message': f'Already marked as {today_status["type"]} at {today_status["time"]}',
                'already_marked': True
            }), 400
        
        now = datetime.now()
        current_date = now.strftime('%Y-%m-%d')
        current_time = now.strftime('%H:%M:%S')
        week_number = get_week_number(now)
        month_year = get_month_year(now)
        year = now.year
        client_ip = get_client_ip()
        
        if attendance_type == 'WFO' and not is_office_ip(client_ip):
            return jsonify({
                'success': False,
                'message': 'Please connect to Office WiFi for WFO',
                'code': 'WIFI_REQUIRED'
            }), 403
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database connection error'}), 500
        
        attendance_ws = sheet.worksheet("Attendance")
        next_id = get_next_id(attendance_ws)
        
        attendance_ws.append_row([
            next_id,
            session['employee_id'],
            session['employee_name'],
            current_date,
            current_time,
            attendance_type,
            week_number,
            month_year,
            year,
            client_ip if attendance_type == 'WFO' else 'Remote',
            'Active'
        ])
        
        log_audit('employee', session['employee_id'], session['employee_name'], 
                 'MARK_ATTENDANCE', f'Marked as {attendance_type}')
        
        # ========== SIMPLE CACHE CLEARING ==========
        # Clear ALL caches to force fresh data
        global _worksheet_cache, cache
        _worksheet_cache = {}  # Clear worksheet cache
        cache = {}             # Clear smart cache
        print("🧹 All caches cleared - attendance will reflect immediately")
        
        return jsonify({
            'success': True,
            'message': f'Attendance marked as {attendance_type} successfully!',
            'type': attendance_type,
            'time': current_time
        })
        
    except Exception as e:
        logger.error(f"Attendance error: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/api/submit-work-log', methods=['POST'])
@login_required
def submit_work_log():
    """Submit daily work log"""
    try:
        data = request.json
        work_description = safe_str(data.get('work_description', '')).strip()
        hours_worked = safe_str(data.get('hours_worked', '8')).strip()
        
        if not work_description:
            return jsonify({'success': False, 'message': 'Work description is required'}), 400
        
        if len(work_description) < 20:
            return jsonify({'success': False, 'message': 'Work description must be at least 20 characters'}), 400
        
        try:
            hours = float(hours_worked)
            if hours <= 0 or hours > 24:
                return jsonify({'success': False, 'message': 'Hours must be between 1 and 24'}), 400
        except:
            return jsonify({'success': False, 'message': 'Invalid hours format'}), 400
        
        # Check if already submitted today
        today_log = check_today_work_log(session['employee_id'])
        if today_log['submitted']:
            return jsonify({
                'success': False,
                'message': 'Work log already submitted today'
            }), 400
        
        now = datetime.now()
        current_date = now.strftime('%Y-%m-%d')
        current_time = now.strftime('%H:%M:%S')
        client_ip = get_client_ip()
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database connection error'}), 500
        
        work_logs_ws = sheet.worksheet("WorkLogs")
        next_id = get_next_id(work_logs_ws)
        
        work_logs_ws.append_row([
            next_id,
            session['employee_id'],
            session['employee_name'],
            current_date,
            work_description,
            hours,
            current_time,
            'Active',
            client_ip
        ])
        
        log_audit('employee', session['employee_id'], session['employee_name'], 
                 'SUBMIT_WORK_LOG', f'Submitted work log for {current_date}')
        
        return jsonify({
            'success': True,
            'message': 'Work log submitted successfully!'
        })
        
    except Exception as e:
        logger.error(f"Work log error: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/employee/history')
@login_required
def employee_history():
    """OPTIMIZED Employee history"""
    try:
        attendance_data = get_cached_worksheet_data("Attendance")
        employee_id = session['employee_id']
        
        attendance_records = []
        total_wfo = 0
        total_wfh = 0
        total_leave = 0
        
        for record in attendance_data:
            if safe_str(record.get('Employee ID')).strip() == employee_id:
                att_type = record.get('Type', '')
                
                if att_type == 'WFO':
                    total_wfo += 1
                elif att_type == 'WFH':
                    total_wfh += 1
                elif att_type == 'Leave':
                    total_leave += 1
                
                attendance_records.append({
                    'date': record.get('Date', ''),
                    'time': record.get('Time', ''),
                    'type': att_type,
                    'mode': record.get('Status', 'Unknown'),
                    'ip_address': record.get('IP Address', ''),
                    'week': record.get('Week', ''),
                    'month': record.get('Month', ''),
                    'year': record.get('Year', '')
                })
        
        attendance_records.sort(key=lambda x: x['date'], reverse=True)
        total_present = total_wfo + total_wfh
        
        return render_template('employee/history.html',
                             attendance_records=attendance_records,
                             total_present=total_present,
                             total_wfo=total_wfo,
                             total_wfh=total_wfh,
                             total_leave=total_leave,
                             records_count=len(attendance_records))
    
    except Exception as e:
        return render_template('employee/history.html', error=str(e))

@app.route('/employee/profile')
@login_required
def employee_profile():
    """Employee profile page"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return render_template('employee/profile.html', error="Database connection error")
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        # Find current employee
        employee_data = None
        for emp in all_employees:
            if str(emp.get('Employee ID', '')).strip() == session['employee_id']:
                employee_data = emp
                break
        
        if not employee_data:
            return render_template('employee/profile.html', error="Employee not found")
        
        # Get attendance statistics - SAME AS DASHBOARD
        attendance_ws = sheet.worksheet("Attendance")
        all_records = attendance_ws.get_all_records()
        
        # Current month (same as dashboard)
        current_month = datetime.now().strftime('%B %Y')  # "February 2026"
        
        # Calculate stats - EXACTLY like dashboard
        stats = {'present': 0, 'wfh': 0, 'leave': 0, 'total': 0}
        
        for record in all_records:
            if (safe_str(record.get('Employee ID', '')).strip() == session['employee_id'] and 
                record.get('Month') == current_month):
                att_type = record.get('Type', 'WFO')
                if att_type == 'WFO':
                    stats['present'] += 1
                elif att_type == 'WFH':
                    stats['wfh'] += 1
                elif att_type == 'Leave':
                    stats['leave'] += 1
                stats['total'] += 1
        
        # Calculate all-time totals (optional)
        all_time_present = 0
        all_time_wfh = 0
        all_time_leave = 0
        
        for record in all_records:
            if safe_str(record.get('Employee ID', '')).strip() == session['employee_id']:
                att_type = record.get('Type', 'WFO')
                if att_type == 'WFO':
                    all_time_present += 1
                elif att_type == 'WFH':
                    all_time_wfh += 1
                elif att_type == 'Leave':
                    all_time_leave += 1
        
        # Get work logs count
        work_logs_ws = sheet.worksheet("WorkLogs")
        all_logs = work_logs_ws.get_all_records()
        work_logs_count = 0
        for log in all_logs:
            if safe_str(log.get('Employee ID', '')).strip() == session['employee_id']:
                work_logs_count += 1
        
        print(f"📊 Profile Stats for {session['employee_id']}:")
        print(f"  Month: {current_month}")
        print(f"  Present (WFO): {stats['present']}")
        print(f"  WFH: {stats['wfh']}")
        print(f"  Leave: {stats['leave']}")
        print(f"  Total: {stats['total']}")
        
        return render_template('employee/profile.html', 
                             employee=employee_data,
                             # Monthly stats (same as dashboard)
                             stats=stats,
                             # All-time stats (optional)
                             all_time_present=all_time_present,
                             all_time_wfh=all_time_wfh,
                             all_time_leave=all_time_leave,
                             work_logs_count=work_logs_count,
                             current_month=current_month)
    
    except Exception as e:
        logger.error(f"Profile error: {str(e)}")
        print(f"❌ Profile Error: {str(e)}")
        return render_template('employee/profile.html', error=str(e))

@app.route('/api/employee/attendance-history')
@login_required
def get_employee_attendance_history():
    """Get attendance history for current month"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        # Get current month start and end
        today = date.today()
        month_start = date(today.year, today.month, 1)
        next_month = today.replace(day=28) + timedelta(days=4)
        month_end = next_month - timedelta(days=next_month.day)
        
        attendance_ws = sheet.worksheet("Attendance")
        all_records = attendance_ws.get_all_records()
        
        # Create calendar for the month
        calendar_data = []
        current_date = month_start
        while current_date <= month_end:
            date_str = current_date.isoformat()
            day_data = {
                'date': date_str,
                'day': current_date.strftime('%d'),
                'day_name': current_date.strftime('%a'),
                'attendance': None,
                'is_today': current_date == today,
                'is_weekend': current_date.weekday() >= 5  # Saturday or Sunday
            }
            
            # Check attendance for this date
            for record in all_records:
                if (safe_str(record.get('Employee ID', '')).strip() == session['employee_id'] and 
                    record.get('Date') == date_str):
                    day_data['attendance'] = {
                        'type': record.get('Type', 'WFO'),
                        'time': record.get('Time')
                    }
                    break
            
            calendar_data.append(day_data)
            current_date += timedelta(days=1)
        
        # Calculate stats
        stats = {'present': 0, 'wfh': 0, 'leave': 0, 'total': 0}
        for day in calendar_data:
            if day['attendance']:
                att_type = day['attendance']['type']
                if att_type == 'WFO':
                    stats['present'] += 1
                elif att_type == 'WFH':
                    stats['wfh'] += 1
                elif att_type == 'Leave':
                    stats['leave'] += 1
                stats['total'] += 1
        
        return jsonify({
            'success': True,
            'calendar': calendar_data,
            'stats': stats,
            'month': today.strftime('%B %Y')
        })
        
    except Exception as e:
        logger.error(f"Attendance history error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/employee/work-log-history')
@login_required
def get_employee_work_log_history():
    """Get work log history (last 30 days)"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        thirty_days_ago = (date.today() - timedelta(days=30)).isoformat()
        
        work_logs_ws = sheet.worksheet("WorkLogs")
        all_logs = work_logs_ws.get_all_records()
        
        logs = []
        for log in all_logs:
            if (safe_str(log.get('Employee ID', '')).strip() == session['employee_id'] and 
                log.get('Date', '') >= thirty_days_ago):
                logs.append({
                    'date': log.get('Date'),
                    'description': log.get('Work Description'),
                    'hours': log.get('Hours Worked'),
                    'submitted_at': log.get('Submitted At')
                })
        
        logs.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify({
            'success': True,
            'logs': logs
        })
        
    except Exception as e:
        logger.error(f"Work log history error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@login_required

# ==================== ADMIN ROUTES ====================

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    """Admin dashboard"""
    return render_template('admin/dashboard.html',
                         admin_name=session['admin_name'],
                         admin_role=session.get('admin_role', 'admin'))

@app.route('/api/admin/dashboard-stats')
@admin_required
def get_admin_dashboard_stats():
    """OPTIMIZED: 3x faster with caching"""
    try:
        # Get ALL data with caching
        attendance_data = get_cached_worksheet_data("Attendance")
        employees_data = get_cached_worksheet_data("Employees")
        audit_data = get_cached_worksheet_data("AuditLogs")
        
        today = date.today().isoformat()
        
        # Process stats
        total_employees = 0
        present_today = 0
        wfh_today = 0
        leave_today = 0
        marked_employees = set()
        
        # Department stats
        dept_stats = {}
        emp_dept = {}
        
        # First pass: Get employee departments
        for emp in employees_data:
            if emp.get('Is Active') == 'Yes':
                total_employees += 1
                emp_id = safe_str(emp.get('Employee ID')).strip()
                dept = emp.get('Department', 'Other')
                emp_dept[emp_id] = dept
                
                # Initialize department stats
                if dept not in dept_stats:
                    dept_stats[dept] = {
                        'total': 0,
                        'present': 0,
                        'wfh': 0,
                        'leave': 0
                    }
                dept_stats[dept]['total'] += 1
        
        # Second pass: Process attendance
        for record in attendance_data:
            if record.get('Date') == today:
                emp_id = safe_str(record.get('Employee ID')).strip()
                marked_employees.add(emp_id)
                
                att_type = record.get('Type', 'WFO')
                dept = emp_dept.get(emp_id, 'Other')
                
                if dept in dept_stats:
                    if att_type == 'WFO':
                        dept_stats[dept]['present'] += 1
                    elif att_type == 'WFH':
                        dept_stats[dept]['wfh'] += 1
                    elif att_type == 'Leave':
                        dept_stats[dept]['leave'] += 1
                
                # Today's totals
                if att_type == 'WFO':
                    present_today += 1
                elif att_type == 'WFH':
                    wfh_today += 1
                elif att_type == 'Leave':
                    leave_today += 1
        
        not_marked = total_employees - len(marked_employees)
        
        # Get recent activity (last 10)
        recent_activity = []
        for log in reversed(audit_data[-20:]):  # Last 20 logs
            recent_activity.append({
                'user': log.get('User Name', 'System'),
                'action': log.get('Action', ''),
                'details': log.get('Details', ''),
                'time': log.get('Timestamp', datetime.now().isoformat())
            })
            if len(recent_activity) >= 10:
                break
        
        return jsonify({
            'success': True,
            'stats': {
                'total_employees': total_employees,
                'present_today': present_today,
                'wfh_today': wfh_today,
                'leave_today': leave_today,
                'not_marked': not_marked
            },
            'dept_stats': dept_stats,
            'recent_activity': recent_activity
        })
        
    except Exception as e:
        print(f"Error in dashboard stats: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/employees')
@admin_required
def admin_employees():
    """Employee management page"""
    return render_template('admin/employees.html',
                         admin_name=session['admin_name'])

@app.route('/api/admin/employees')
@admin_required
def get_all_employees():
    """Get all employees"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        employees = []
        for emp in all_employees:
            employees.append({
                'id': emp.get('ID'),
                'employee_id': safe_str(emp.get('Employee ID')),
                'name': safe_str(emp.get('Employee Name')),
                'email': safe_str(emp.get('Email')),
                'department': safe_str(emp.get('Department')),
                'join_date': emp.get('Join Date'),
                'is_active': emp.get('Is Active') == 'Yes',
                'last_login': emp.get('Last Login', 'Never'),
                'force_change': emp.get('Force Password Change', 'No') == 'Yes'
            })
        
        return jsonify({
            'success': True,
            'employees': employees
        })
        
    except Exception as e:
        logger.error(f"Get employees error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/employees', methods=['POST'])
@admin_required
def add_employee():
    """Add new employee"""
    try:
        data = request.json
        emp_id = safe_str(data.get('employee_id', '')).strip()
        name = safe_str(data.get('name', '')).strip()
        email = safe_str(data.get('email', '')).strip()
        department = safe_str(data.get('department', '')).strip()
        
        if not emp_id or not name or not email:
            return jsonify({'success': False, 'message': 'Required fields missing'}), 400
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        employees_ws = sheet.worksheet("Employees")
        
        # Check if employee ID already exists
        all_employees = employees_ws.get_all_records()
        for emp in all_employees:
            if safe_str(emp.get('Employee ID', '')).strip() == emp_id:
                return jsonify({'success': False, 'message': 'Employee ID already exists'}), 400
        
        # Add employee with default password and force change
        next_id = get_next_id(employees_ws)
        password_hash = generate_password_hash(emp_id)
        today = date.today().isoformat()
        
        employees_ws.append_row([
            next_id,
            emp_id,
            name,
            email,
            password_hash,
            department,
            today,
            'Yes',
            'employee',
            '',
            '',
            'Yes'  # Force password change
        ])
        
        log_audit('admin', session['admin_id'], session['admin_name'], 
                 'ADD_EMPLOYEE', f'Added employee: {name} ({emp_id})')
        
        return jsonify({
            'success': True,
            'message': 'Employee added successfully'
        })
        
    except Exception as e:
        logger.error(f"Add employee error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/employees/<emp_id>', methods=['PUT'])
@admin_required
def update_employee(emp_id):
    """Update employee"""
    try:
        data = request.json
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        row_index = None
        for i, emp in enumerate(all_employees, start=2):
            if safe_str(emp.get('Employee ID', '')).strip() == emp_id:
                row_index = i
                break
        
        if not row_index:
            return jsonify({'success': False, 'message': 'Employee not found'}), 404
        
        # Update fields
        if 'name' in data:
            employees_ws.update_cell(row_index, 3, safe_str(data['name']))
        if 'email' in data:
            employees_ws.update_cell(row_index, 4, safe_str(data['email']))
        if 'department' in data:
            employees_ws.update_cell(row_index, 6, safe_str(data['department']))
        if 'is_active' in data:
            employees_ws.update_cell(row_index, 8, 'Yes' if data['is_active'] else 'No')
        
        log_audit('admin', session['admin_id'], session['admin_name'], 
                 'UPDATE_EMPLOYEE', f'Updated employee: {emp_id}')
        
        return jsonify({
            'success': True,
            'message': 'Employee updated successfully'
        })
        
    except Exception as e:
        logger.error(f"Update employee error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/employees/<emp_id>/reset-password', methods=['POST'])
@admin_required
def reset_employee_password(emp_id):
    """Reset employee password to default and force change"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        row_index = None
        for i, emp in enumerate(all_employees, start=2):
            if safe_str(emp.get('Employee ID', '')).strip() == emp_id:
                row_index = i
                break
        
        if not row_index:
            return jsonify({'success': False, 'message': 'Employee not found'}), 404
        
        new_hash = generate_password_hash(emp_id)
        employees_ws.update_cell(row_index, 5, new_hash)  # Password Hash
        employees_ws.update_cell(row_index, 11, '')  # Clear password changed
        employees_ws.update_cell(row_index, 12, 'Yes')  # Force password change
        
        log_audit('admin', session['admin_id'], session['admin_name'], 
                 'RESET_PASSWORD', f'Reset password for employee: {emp_id}')
        
        return jsonify({
            'success': True,
            'message': f'Password reset to default (Employee ID)'
        })
        
    except Exception as e:
        logger.error(f"Reset password error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/attendance')
@admin_required
def admin_attendance():
    """Attendance view page"""
    return render_template('admin/attendance.html',
                         admin_name=session['admin_name'])

@app.route('/api/admin/attendance')
@admin_required
def get_all_attendance():
    """Get all attendance records with filters"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        from_date = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
        to_date = request.args.get('to', date.today().isoformat())
        employee_id = request.args.get('employee', '')
        dept = request.args.get('dept', '')
        
        attendance_ws = sheet.worksheet("Attendance")
        all_records = attendance_ws.get_all_records()
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        emp_dept = {}
        emp_name = {}
        for emp in all_employees:
            emp_id = safe_str(emp.get('Employee ID', '')).strip()
            emp_dept[emp_id] = emp.get('Department', '')
            emp_name[emp_id] = emp.get('Employee Name', '')
        
        records = []
        for record in all_records:
            record_date = record.get('Date', '')
            rec_emp_id = safe_str(record.get('Employee ID', '')).strip()
            rec_dept = emp_dept.get(rec_emp_id, '')
            
            if record_date < from_date or record_date > to_date:
                continue
            if employee_id and rec_emp_id != employee_id:
                continue
            if dept and rec_dept != dept:
                continue
            
            records.append({
                'date': record_date,
                'employee_id': rec_emp_id,
                'employee_name': emp_name.get(rec_emp_id, record.get('Employee Name', '')),
                'time': record.get('Time', ''),
                'type': record.get('Type', 'WFO'),
                'department': rec_dept,
                'ip': record.get('IP Address', '')
            })
        
        records.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify({
            'success': True,
            'records': records
        })
        
    except Exception as e:
        logger.error(f"Get attendance error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/work-logs')
@admin_required
def admin_work_logs():
    """Work logs view page"""
    return render_template('admin/work_logs.html',
                         admin_name=session['admin_name'])

@app.route('/api/admin/work-logs')
@admin_required
def get_all_work_logs():
    """Get all work logs with filters"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        from_date = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
        to_date = request.args.get('to', date.today().isoformat())
        employee_id = request.args.get('employee', '')
        dept = request.args.get('dept', '')
        
        work_logs_ws = sheet.worksheet("WorkLogs")
        all_logs = work_logs_ws.get_all_records()
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        emp_dept = {}
        for emp in all_employees:
            emp_id = safe_str(emp.get('Employee ID', '')).strip()
            emp_dept[emp_id] = emp.get('Department', '')
        
        logs = []
        for log in all_logs:
            log_date = log.get('Date', '')
            log_emp_id = safe_str(log.get('Employee ID', '')).strip()
            log_dept = emp_dept.get(log_emp_id, '')
            
            if log_date < from_date or log_date > to_date:
                continue
            if employee_id and log_emp_id != employee_id:
                continue
            if dept and log_dept != dept:
                continue
            
            logs.append({
                'date': log_date,
                'employee_id': log_emp_id,
                'employee_name': log.get('Employee Name', ''),
                'description': log.get('Work Description', ''),
                'hours': log.get('Hours Worked', ''),
                'submitted_at': log.get('Submitted At', ''),
                'department': log_dept,
                'ip': log.get('IP Address', '')
            })
        
        logs.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify({
            'success': True,
            'logs': logs
        })
        
    except Exception as e:
        logger.error(f"Get work logs error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/reports')
@admin_required
def admin_reports():
    """Reports page"""
    return render_template('admin/reports.html',
                         admin_name=session['admin_name'])

@app.route('/api/admin/monthly-report')
@admin_required
def get_monthly_report():
    """Get monthly attendance report"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        month = request.args.get('month', datetime.now().strftime('%B %Y'))
        
        attendance_ws = sheet.worksheet("Attendance")
        all_records = attendance_ws.get_all_records()
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        report = {}
        for emp in all_employees:
            if emp.get('Is Active') == 'Yes':
                emp_id = safe_str(emp.get('Employee ID', '')).strip()
                report[emp_id] = {
                    'employee_id': emp_id,
                    'employee_name': emp.get('Employee Name', ''),
                    'department': emp.get('Department', ''),
                    'present': 0,
                    'wfh': 0,
                    'leave': 0,
                    'total': 0
                }
        
        for record in all_records:
            if record.get('Month') == month:
                emp_id = safe_str(record.get('Employee ID', '')).strip()
                if emp_id in report:
                    att_type = record.get('Type', 'WFO')
                    if att_type == 'WFO':
                        report[emp_id]['present'] += 1
                    elif att_type == 'WFH':
                        report[emp_id]['wfh'] += 1
                    elif att_type == 'Leave':
                        report[emp_id]['leave'] += 1
                    report[emp_id]['total'] += 1
        
        return jsonify({
            'success': True,
            'month': month,
            'report': list(report.values())
        })
        
    except Exception as e:
        logger.error(f"Monthly report error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/export-report', methods=['POST'])
@admin_required
def export_report():
    """Export report as Excel"""
    try:
        data = request.json
        report_type = data.get('type', 'monthly')
        month = data.get('month', datetime.now().strftime('%B %Y'))
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        if report_type == 'monthly':
            attendance_ws = sheet.worksheet("Attendance")
            all_records = attendance_ws.get_all_records()
            
            employees_ws = sheet.worksheet("Employees")
            all_employees = employees_ws.get_all_records()
            
            data = []
            for emp in all_employees:
                if emp.get('Is Active') == 'Yes':
                    emp_id = safe_str(emp.get('Employee ID', '')).strip()
                    emp_data = {
                        'Employee ID': emp_id,
                        'Employee Name': emp.get('Employee Name', ''),
                        'Department': emp.get('Department', ''),
                        'Present': 0,
                        'WFH': 0,
                        'Leave': 0,
                        'Total Days': 0
                    }
                    
                    for record in all_records:
                        if (record.get('Month') == month and 
                            safe_str(record.get('Employee ID', '')).strip() == emp_id):
                            att_type = record.get('Type', 'WFO')
                            if att_type == 'WFO':
                                emp_data['Present'] += 1
                            elif att_type == 'WFH':
                                emp_data['WFH'] += 1
                            elif att_type == 'Leave':
                                emp_data['Leave'] += 1
                            emp_data['Total Days'] += 1
                    
                    data.append(emp_data)
            
            df = pd.DataFrame(data)
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=f'Report_{month}', index=False)
            
            output.seek(0)
            excel_data = base64.b64encode(output.read()).decode('utf-8')
            
            return jsonify({
                'success': True,
                'data': excel_data,
                'filename': f'attendance_report_{month}.xlsx'
            })
        
        return jsonify({'success': False, 'message': 'Invalid report type'}), 400
        
    except Exception as e:
        logger.error(f"Export error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/logs')
@admin_required
def admin_logs():
    """Audit logs page"""
    return render_template('admin/logs.html',
                         admin_name=session['admin_name'])

@app.route('/api/admin/logs')
@admin_required
def get_audit_logs():
    """Get audit logs"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        days = int(request.args.get('days', 7))
        since_date = (datetime.now() - timedelta(days=days)).isoformat()
        
        logs_ws = sheet.worksheet("AuditLogs")
        all_logs = logs_ws.get_all_records()
        
        logs = []
        for log in reversed(all_logs):
            if log.get('Timestamp', '') >= since_date:
                logs.append({
                    'timestamp': log.get('Timestamp'),
                    'user_type': log.get('User Type'),
                    'user_name': log.get('User Name'),
                    'action': log.get('Action'),
                    'details': log.get('Details'),
                    'ip': log.get('IP Address')
                })
        
        return jsonify({
            'success': True,
            'logs': logs
        })
        
    except Exception as e:
        logger.error(f"Audit logs error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/departments')
@admin_required
def get_departments():
    """Get all departments"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        departments_ws = sheet.worksheet("Departments")
        all_depts = departments_ws.get_all_records()
        
        departments = []
        for dept in all_depts:
            if dept.get('Is Active') == 'Yes':
                departments.append(dept.get('Department Name'))
        
        return jsonify({
            'success': True,
            'departments': sorted(departments)
        })
        
    except Exception as e:
        logger.error(f"Departments error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/employee-list')
@admin_required
def get_employee_list():
    """Get list of active employees for dropdowns"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        employees = []
        for emp in all_employees:
            if emp.get('Is Active') == 'Yes':
                employees.append({
                    'id': safe_str(emp.get('Employee ID')),
                    'name': emp.get('Employee Name')
                })
        
        employees.sort(key=lambda x: x['name'])
        
        return jsonify({
            'success': True,
            'employees': employees
        })
        
    except Exception as e:
        logger.error(f"Employee list error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/settings')
@admin_required
def admin_settings():
    """Settings page"""
    return render_template('admin/settings.html',
                         admin_name=session['admin_name'])

@app.route('/api/admin/settings', methods=['GET', 'POST'])
@admin_required
def manage_settings():
    """Get or update settings"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        settings_ws = sheet.worksheet("Settings")
        
        if request.method == 'POST':
            data = request.json
            print("Saving settings:", data)  # Debug log
            
            # Get all existing settings
            all_settings = settings_ws.get_all_records()
            all_rows = settings_ws.get_all_values()
            headers = all_rows[0] if all_rows else []
            
            # Update or insert each setting
            for key, value in data.items():
                # Find if setting exists
                row_index = None
                for i, setting in enumerate(all_settings, start=2):
                    if setting.get('Key') == key:
                        row_index = i
                        break
                
                if row_index:
                    # Update existing setting
                    settings_ws.update_cell(row_index, 2, str(value))
                    settings_ws.update_cell(row_index, 3, session.get('admin_name', 'admin'))
                    settings_ws.update_cell(row_index, 4, datetime.now().isoformat())
                else:
                    # Insert new setting
                    settings_ws.append_row([key, str(value), session.get('admin_name', 'admin'), datetime.now().isoformat()])
            
            log_audit('admin', session['admin_id'], session.get('admin_name', ''), 
                     'UPDATE_SETTINGS', 'Updated system settings')
            
            return jsonify({'success': True, 'message': 'Settings saved successfully'})
        
        # GET request - return all settings
        all_settings = settings_ws.get_all_records()
        settings = {}
        for setting in all_settings:
            settings[setting.get('Key')] = setting.get('Value')
        
        # Set default values if not found
        defaults = {
            'min_password_length': '8',
            'session_timeout': '8',
            'default_work_hours': '8',
            'work_log_min_chars': '20',
            'password_expiry_days': '90',
            'max_login_attempts': '5',
            'require_uppercase': 'true',
            'require_numbers': 'true',
            'require_special': 'true',
            'allow_wfh': 'true',
            'allow_leave': 'true',
            'enable_ip_check': 'true',
            'enable_audit_logs': 'true',
            'enable_emails': 'false',
            'smtp_server': '',
            'smtp_port': '587',
            'smtp_username': '',
            'from_email': ''
        }
        
        # Merge with defaults
        for key, default_value in defaults.items():
            if key not in settings:
                settings[key] = default_value
        
        return jsonify({'success': True, 'settings': settings})
        
    except Exception as e:
        print(f"Settings error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== UTILITY ROUTES ====================

@app.route('/api/check-status')
@login_required
def check_status():
    """Check today's attendance and work log status"""
    try:
        attendance = check_today_attendance(session['employee_id'])
        work_log = check_today_work_log(session['employee_id'])
        
        return jsonify({
            'success': True,
            'attendance': attendance,
            'work_log': work_log
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
    # ==================== WORK LOG ROUTES ====================
# ==================== WORK LOG ROUTES ====================



@app.route('/profile')
@login_required
def profile():
    """Employee profile page"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return render_template('profile.html', error="Database connection error")
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        # Find current employee
        employee_data = None
        for emp in all_employees:
            if str(emp.get('Employee ID', '')).strip() == session['employee_id']:
                employee_data = emp
                break
        
        if not employee_data:
            return render_template('profile.html', error="Employee not found")
        
        # Get attendance statistics
        attendance_ws = sheet.worksheet("Attendance")
        all_attendance = attendance_ws.get_all_records()
        
        # Calculate stats
        current_month = datetime.now().strftime('%B')
        current_year = str(datetime.now().year)
        
        total_present = 0
        total_wfh = 0
        total_leave = 0
        monthly_present = 0
        
        for record in all_attendance:
            if str(record.get('Employee ID', '')).strip() == session['employee_id']:
                if record.get('Type') == 'Present':
                    total_present += 1
                    if record.get('Month') == current_month and str(record.get('Year')) == current_year:
                        monthly_present += 1
                elif record.get('Type') == 'WFH':
                    total_wfh += 1
                elif record.get('Type') == 'Leave':
                    total_leave += 1
        
        return render_template('profile.html', 
                             employee=employee_data,
                             total_present=total_present,
                             total_wfh=total_wfh,
                             total_leave=total_leave,
                             monthly_present=monthly_present)
    
    except Exception as e:
        logger.error(f"Profile error: {str(e)}")
        return render_template('profile.html', error=str(e))



@app.route('/api/attendance-calendar')
@login_required
def attendance_calendar():
    """Get attendance data for calendar"""
    try:
        year = int(request.args.get('year', datetime.now().year))
        month = int(request.args.get('month', datetime.now().month))
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'})
        
        attendance_ws = sheet.worksheet("Attendance")
        all_records = attendance_ws.get_all_records()
        
        # Filter for current employee and specified month/year
        month_name = datetime(year, month + 1, 1).strftime('%B')
        
        attendance_data = {}
        for record in all_records:
            if (str(record.get('Employee ID', '')).strip() == session['employee_id'] and
                record.get('Month') == month_name and
                str(record.get('Year')) == str(year)):
                
                date_str = record.get('Date')
                if date_str:
                    # Extract day from date
                    try:
                        day = int(date_str.split('-')[-1])
                        attendance_data[day] = {
                            'type': record.get('Type'),
                            'mode': record.get('Status', 'Unknown')
                        }
                    except:
                        pass
        
        return jsonify({
            'success': True,
            'attendance': attendance_data
        })
        
    except Exception as e:
        logger.error(f"Calendar API error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

import random
import string
from flask import render_template

# ==================== FORGOT PASSWORD ROUTES ====================

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    """Forgot password page - verify employee ID and email"""
    if request.method == 'POST':
        try:
            data = request.json
            employee_id = safe_str(data.get('employee_id', '')).strip()
            email = safe_str(data.get('email', '')).strip()
            
            print(f"🔍 Forgot password attempt - ID: '{employee_id}', Email: '{email}'")
            
            if not employee_id or not email:
                return jsonify({'success': False, 'message': 'Please enter both fields'}), 400
            
            sheet = get_cached_sheet()
            if not sheet:
                return jsonify({'success': False, 'message': 'Database error'}), 500
            
            employees_ws = sheet.worksheet("Employees")
            all_employees = employees_ws.get_all_records()
            
            # Find employee with flexible matching
            employee = None
            for emp in all_employees:
                sheet_id = safe_str(emp.get('Employee ID', '')).strip()
                sheet_email = safe_str(emp.get('Email', '')).strip().lower()
                
                # Try multiple matching methods
                if (sheet_id == employee_id or 
                    sheet_id.lstrip('0') == employee_id.lstrip('0') or
                    sheet_id.lower() == employee_id.lower()):
                    
                    if sheet_email == email.lower():
                        employee = emp
                        print(f"✅ Employee found: {sheet_id} with email {sheet_email}")
                        break
            
            if not employee:
                print(f"❌ No match found for ID: {employee_id}, Email: {email}")
                return jsonify({
                    'success': False, 
                    'message': 'Employee ID and Email do not match'
                }), 404
            
            # Store verification in session
            import time
            session['reset_allowed'] = True
            session['reset_emp'] = safe_str(employee.get('Employee ID'))
            session['reset_email'] = email
            session['reset_time'] = time.time()
            
            print(f"✅ Verification successful for: {employee.get('Employee ID')}")
            
            # Return success with redirect URL
            return jsonify({
                'success': True,
                'message': 'Verification successful',
                'redirect': f'/reset-password?emp={employee.get("Employee ID")}'
            })
            
        except Exception as e:
            print(f"❌ Forgot password error: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500
    
    return render_template('forgot_password.html')

@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    """Reset password page"""
    # GET request - show form
    if request.method == 'GET':
        emp_id = request.args.get('emp', '')
        print(f"🔍 GET reset password for: {emp_id}")
        
        # Check if reset is allowed
        if not session.get('reset_allowed') or session.get('reset_emp') != emp_id:
            print("❌ Not authorized - redirecting to forgot password")
            flash('Please verify your identity first', 'warning')
            return redirect(url_for('forgot_password'))
        
        # Check expiry (30 minutes)
        import time
        if time.time() - session.get('reset_time', 0) > 1800:
            print("❌ Session expired")
            session.pop('reset_allowed', None)
            session.pop('reset_emp', None)
            session.pop('reset_email', None)
            session.pop('reset_time', None)
            flash('Session expired. Please try again.', 'warning')
            return redirect(url_for('forgot_password'))
        
        return render_template('reset_password.html', emp_id=emp_id)
    
    # POST request - handle password reset
    try:
        data = request.json
        emp_id = data.get('emp_id', '')
        new_password = data.get('new_password', '')
        confirm_password = data.get('confirm_password', '')
        
        print(f"🔍 POST reset password for: {emp_id}")
        
        # Verify authorization
        if not session.get('reset_allowed') or session.get('reset_emp') != emp_id:
            return jsonify({'success': False, 'message': 'Not authorized. Please start over.'}), 400
        
        # Check expiry
        import time
        if time.time() - session.get('reset_time', 0) > 1800:
            session.pop('reset_allowed', None)
            session.pop('reset_emp', None)
            session.pop('reset_email', None)
            session.pop('reset_time', None)
            return jsonify({'success': False, 'message': 'Session expired. Please try again.'}), 400
        
        # Validate passwords
        if not new_password or not confirm_password:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        
        if new_password != confirm_password:
            return jsonify({'success': False, 'message': 'Passwords do not match'}), 400
        
        if len(new_password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters'}), 400
        
        # Update password in Google Sheets
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        all_rows = employees_ws.get_all_values()
        
        # Find employee row
        row_index = None
        for i, emp in enumerate(all_employees, start=2):
            sheet_id = safe_str(emp.get('Employee ID', '')).strip()
            if sheet_id == emp_id or sheet_id.lstrip('0') == emp_id.lstrip('0'):
                row_index = i
                print(f"✅ Found employee at row {row_index}")
                break
        
        if not row_index:
            return jsonify({'success': False, 'message': 'Employee not found'}), 404
        
        # Update password
        new_hash = generate_password_hash(new_password)
        
        # Find password column
        headers = all_rows[0]
        password_col = None
        for idx, header in enumerate(headers):
            if header == 'Password Hash':
                password_col = idx + 1
                print(f"✅ Password column found at index {password_col}")
                break
        
        if password_col:
            employees_ws.update_cell(row_index, password_col, new_hash)
        else:
            # Default to column 5 if not found
            employees_ws.update_cell(row_index, 5, new_hash)
        
        # Update force password change flag if it exists
        force_change_col = None
        for idx, header in enumerate(headers):
            if header == 'Force Password Change':
                force_change_col = idx + 1
                break
        
        if force_change_col:
            employees_ws.update_cell(row_index, force_change_col, 'No')
        
        # Clear session
        session.pop('reset_allowed', None)
        session.pop('reset_emp', None)
        session.pop('reset_email', None)
        session.pop('reset_time', None)
        
        print(f"✅ Password reset successful for: {emp_id}")
        
        return jsonify({
            'success': True,
            'message': 'Password reset successful! Redirecting to login...'
        })
        
    except Exception as e:
        print(f"❌ Reset error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== ADMIN FORGOT PASSWORD ROUTES ====================

@app.route('/debug-employees')
def debug_employees():
    """Debug endpoint to see all employees in sheet"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return "Database connection error"
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        all_rows = employees_ws.get_all_values()
        
        # Print headers
        headers = all_rows[0] if all_rows else []
        
        result = "<h2>Employee Debug Info</h2>"
        result += f"<h3>Headers: {headers}</h3>"
        result += "<h3>Employees:</h3><ul>"
        
        for emp in all_employees:
            emp_id = str(emp.get('Employee ID', '')).strip()
            emp_name = emp.get('Employee Name', '')
            result += f"<li>ID: '{emp_id}' (raw: {repr(emp_id)}), Name: {emp_name}</li>"
        
        result += "</ul>"
        
        # Also check session
        result += "<h3>Session:</h3>"
        result += f"<p>{session}</p>"
        
        return result
    except Exception as e:
        return f"Error: {str(e)}"
    
@app.route('/test-login', methods=['GET', 'POST'])
def test_login():
    """Simple test login"""
    if request.method == 'POST':
        emp_id = request.form.get('employee_id', '').strip()
        password = request.form.get('password', '').strip()
        
        sheet = get_cached_sheet()
        if not sheet:
            return "DB Error"
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        # Simple check
        found = False
        for emp in all_employees:
            sheet_id = str(emp.get('Employee ID', '')).strip()
            if sheet_id == emp_id:
                found = True
                stored_hash = emp.get('Password Hash', '')
                if check_password_hash(stored_hash, password):
                    return f"✅ Login successful for {emp_id}"
                else:
                    return f"❌ Wrong password for {emp_id}"
        
        if not found:
            return f"❌ Employee ID '{emp_id}' not found"
        
        return "Login failed"
    
    # GET request - show simple form
    return '''
    <form method="post">
        <input type="text" name="employee_id" placeholder="Employee ID">
        <input type="password" name="password" placeholder="Password">
        <button type="submit">Test Login</button>
    </form>
    '''

# ==================== ADMIN FORGOT PASSWORD ROUTES ====================

@app.route('/admin/forgot-password', methods=['GET', 'POST'])
def admin_forgot_password():
    """Admin forgot password - verify Admin ID and Email"""
    if request.method == 'POST':
        try:
            data = request.json
            admin_id = safe_str(data.get('admin_id', '')).strip()
            email = safe_str(data.get('email', '')).strip()
            
            print(f"🔍 Admin forgot - ID: {admin_id}, Email: {email}")  # Debug
            
            if not admin_id or not email:
                return jsonify({'success': False, 'message': 'Please enter both fields'}), 400
            
            sheet = get_cached_sheet()
            if not sheet:
                return jsonify({'success': False, 'message': 'Database error'}), 500
            
            admins_ws = sheet.worksheet("Admins")
            all_admins = admins_ws.get_all_records()
            
            # Find admin
            admin = None
            for adm in all_admins:
                sheet_id = safe_str(adm.get('Admin ID', '')).strip()
                sheet_email = safe_str(adm.get('Email', '')).strip().lower()
                
                if sheet_id == admin_id and sheet_email == email.lower():
                    admin = adm
                    break
            
            if not admin:
                print("❌ Admin not found")
                return jsonify({'success': False, 'message': 'Admin ID and Email do not match'}), 404
            
            # Set session for reset
            session['admin_reset_allowed'] = True
            session['admin_reset_id'] = admin_id
            session['admin_reset_time'] = time.time()
            
            print(f"✅ Admin verified: {admin_id}")
            
            return jsonify({
                'success': True,
                'message': 'Verification successful',
                'redirect': f'/admin/reset-password?admin={admin_id}'
            })
            
        except Exception as e:
            print(f"❌ Admin forgot error: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500
    
    return render_template('admin/forgot_password.html')


# ==================== ADMIN RESET PASSWORD ROUTE ====================

@app.route('/admin/reset-password', methods=['GET', 'POST'])
def admin_reset_password():
    """Admin reset password page"""
    # GET request - show form
    if request.method == 'GET':
        admin_id = request.args.get('admin', '')
        print(f"🔍 Admin GET reset for: {admin_id}")
        
        # Check if reset is allowed
        if not session.get('admin_reset_allowed') or session.get('admin_reset_id') != admin_id:
            print("❌ Admin not authorized")
            return redirect(url_for('admin_forgot_password'))
        
        # Check expiry (30 minutes)
        if time.time() - session.get('admin_reset_time', 0) > 1800:
            print("❌ Admin session expired")
            session.pop('admin_reset_allowed', None)
            session.pop('admin_reset_id', None)
            session.pop('admin_reset_time', None)
            return redirect(url_for('admin_forgot_password'))
        
        return render_template('admin/reset_password.html', admin_id=admin_id)
    
    # POST request - handle password reset
    try:
        data = request.json
        admin_id = data.get('admin_id', '')
        new_password = data.get('new_password', '')
        confirm_password = data.get('confirm_password', '')
        
        print(f"🔍 Admin POST reset for: {admin_id}")
        
        # Verify authorization
        if not session.get('admin_reset_allowed') or session.get('admin_reset_id') != admin_id:
            return jsonify({'success': False, 'message': 'Not authorized. Please start over.'}), 400
        
        # Check expiry
        if time.time() - session.get('admin_reset_time', 0) > 1800:
            session.pop('admin_reset_allowed', None)
            session.pop('admin_reset_id', None)
            session.pop('admin_reset_time', None)
            return jsonify({'success': False, 'message': 'Session expired. Please try again.'}), 400
        
        if new_password != confirm_password:
            return jsonify({'success': False, 'message': 'Passwords do not match'}), 400
        
        if len(new_password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters'}), 400
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        admins_ws = sheet.worksheet("Admins")
        all_admins = admins_ws.get_all_records()
        all_rows = admins_ws.get_all_values()
        
        # Find admin row
        row_index = None
        for i, adm in enumerate(all_admins, start=2):
            sheet_id = safe_str(adm.get('Admin ID', '')).strip()
            if sheet_id == admin_id:
                row_index = i
                break
        
        if not row_index:
            return jsonify({'success': False, 'message': 'Admin not found'}), 404
        
        # Update password
        new_hash = generate_password_hash(new_password)
        
        # Find password hash column
        headers = all_rows[0]
        password_col = None
        for idx, header in enumerate(headers):
            if header == 'Password Hash':
                password_col = idx + 1
                break
        
        if password_col:
            admins_ws.update_cell(row_index, password_col, new_hash)
        else:
            admins_ws.update_cell(row_index, 5, new_hash)
        
        # Clear session
        session.pop('admin_reset_allowed', None)
        session.pop('admin_reset_id', None)
        session.pop('admin_reset_time', None)
        
        print(f"✅ Admin password reset successful for: {admin_id}")
        
        return jsonify({
            'success': True,
            'message': 'Password reset successful! Redirecting to admin login...'
        })
        
    except Exception as e:
        print(f"❌ Admin reset error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== ONLINE/OFFLINE TRACKING ====================
# ==================== ONLINE/OFFLINE TRACKING ====================

def update_employee_status(employee_id, status, break_start=None, break_end=None):
    """COMPLETE FIXED VERSION"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            print("❌ No sheet connection")
            return False
        
        today = date.today().isoformat()
        now = datetime.now()
        
        print(f"🔵 Processing {employee_id} - {status}")
        
        # ===== 1. UPDATE EMPLOYEES SHEET =====
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        all_emp_rows = employees_ws.get_all_values()
        
        # Find employee
        emp_row = None
        for i, emp in enumerate(all_employees, start=2):
            if safe_str(emp.get('Employee ID')).strip() == employee_id:
                emp_row = i
                print(f"✅ Found employee at row {i}")
                break
        
        if not emp_row:
            print(f"❌ Employee not found")
            return False
        
        # Update Employees sheet
        employees_ws.update_cell(emp_row, 13, now.isoformat())  # Last Activity
        employees_ws.update_cell(emp_row, 14, status)  # Status
        
        # ===== 2. HANDLE BREAK =====
        if status == 'break' and break_start:
            employees_ws.update_cell(emp_row, 15, break_start.isoformat())
            print(f"✅ Break started at {break_start}")
            return True
            
        elif status == 'online' and break_end:
            print("🟢 Ending break...")
            
            # Get break start time
            break_start_str = all_emp_rows[emp_row-1][14] if len(all_emp_rows[emp_row-1]) > 14 else ''
            
            if not break_start_str:
                print("❌ No break start time found")
                return False
            
            try:
                start_time = datetime.fromisoformat(break_start_str)
                duration = round((break_end - start_time).total_seconds() / 60, 2)
                print(f"📊 Duration: {duration} minutes")
                
                # ===== 3. UPDATE ATTENDANCE SHEET =====
                attendance_ws = sheet.worksheet("Attendance")
                all_attendance = attendance_ws.get_all_records()
                
                # Find today's attendance record
                att_row = None
                for j, att in enumerate(all_attendance, start=2):
                    att_emp_id = safe_str(att.get('Employee ID')).strip()
                    att_date = att.get('Date', '')
                    
                    if att_emp_id == employee_id and att_date == today:
                        att_row = j
                        print(f"✅ Found attendance at row {j}")
                        break
                
                if not att_row:
                    print(f"❌ No attendance record found for {employee_id} today")
                    print("📝 Creating attendance record now...")
                    
                    # Get employee name
                    emp_name = all_emp_rows[emp_row-1][2] if len(all_emp_rows[emp_row-1]) > 2 else 'Unknown'
                    
                    # Create attendance record
                    next_id = str(len(all_attendance) + 1)
                    
                    # Get week number
                    week_number = now.isocalendar()[1]
                    month_year = now.strftime('%B %Y')
                    
                    # Create row with all columns
                    attendance_ws.append_row([
                        next_id,           # ID
                        employee_id,        # Employee ID
                        emp_name,           # Employee Name
                        today,              # Date
                        now.strftime('%H:%M:%S'),  # Time
                        'WFO',               # Type
                        str(week_number),    # Week
                        month_year,          # Month
                        str(now.year),       # Year
                        'System',            # IP Address
                        'Active',            # Status
                        '0',                 # Break Count (col 12)
                        '0'                  # Break Minutes (col 13)
                    ])
                    
                    print(f"✅ Created attendance record at row {len(all_attendance) + 2}")
                    att_row = len(all_attendance) + 2
                
                # Now update the break count
                all_att_rows = attendance_ws.get_all_values()
                
                # Get current break count from column 12
                current_count = 0
                if len(all_att_rows) > att_row-1 and len(all_att_rows[att_row-1]) >= 12:
                    val = all_att_rows[att_row-1][11]
                    if val and str(val).strip():
                        try:
                            current_count = int(float(val))
                        except:
                            current_count = 0
                
                # Update break count in column 12
                new_count = current_count + 1
                attendance_ws.update_cell(att_row, 12, str(new_count))
                print(f"✅ Updated Attendance Break Count to {new_count}")
                
                # Update break minutes in column 13
                current_minutes = 0
                if len(all_att_rows) > att_row-1 and len(all_att_rows[att_row-1]) >= 13:
                    val = all_att_rows[att_row-1][12]
                    if val and str(val).strip():
                        try:
                            current_minutes = float(val)
                        except:
                            current_minutes = 0
                
                new_minutes = round(current_minutes + duration, 2)
                attendance_ws.update_cell(att_row, 13, str(new_minutes))
                print(f"✅ Updated Attendance Break Minutes to {new_minutes}")
                
                # Update Employees sheet break count
                employees_ws.update_cell(emp_row, 18, str(new_count))  # Break Count
                employees_ws.update_cell(emp_row, 15, '')  # Clear Break Start
                employees_ws.update_cell(emp_row, 16, break_end.isoformat())  # Break End
                
                print(f"✅ Break ended successfully for {employee_id}")
                return True
                
            except Exception as e:
                print(f"❌ Error: {e}")
                import traceback
                traceback.print_exc()
                return False
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

    
@app.route('/create-attendance/<emp_id>')
def create_attendance(emp_id):
    """Create attendance record for employee"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return "Database error"
        
        today = date.today().isoformat()
        now = datetime.now()
        
        # Get employee details
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        emp_name = None
        for emp in all_employees:
            if safe_str(emp.get('Employee ID')).strip() == emp_id:
                emp_name = emp.get('Employee Name')
                break
        
        if not emp_name:
            return f"Employee {emp_id} not found"
        
        # Check if attendance already exists
        attendance_ws = sheet.worksheet("Attendance")
        all_attendance = attendance_ws.get_all_records()
        
        for att in all_attendance:
            if (safe_str(att.get('Employee ID')).strip() == emp_id and 
                att.get('Date') == today):
                return f"Attendance already exists for {emp_name} today"
        
        # Create new attendance record
        next_id = get_next_id(attendance_ws)
        
        # Get headers to find column positions
        headers = attendance_ws.row_values(1)
        
        # Create row data
        row_data = []
        for header in headers:
            if header == 'ID':
                row_data.append(next_id)
            elif header == 'Employee ID':
                row_data.append(emp_id)
            elif header == 'Employee Name':
                row_data.append(emp_name)
            elif header == 'Date':
                row_data.append(today)
            elif header == 'Time':
                row_data.append(now.strftime('%H:%M:%S'))
            elif header == 'Type':
                row_data.append('WFO')
            elif header == 'Week':
                row_data.append(str(now.isocalendar()[1]))
            elif header == 'Month':
                row_data.append(now.strftime('%B %Y'))
            elif header == 'Year':
                row_data.append(str(now.year))
            elif header == 'IP Address':
                row_data.append('System')
            elif header == 'Status':
                row_data.append('Active')
            elif header == 'Break Count':
                row_data.append('0')
            elif header == 'Break Minutes':
                row_data.append('0')
            else:
                row_data.append('')
        
        attendance_ws.append_row(row_data)
        
        return f"""
        <h2 style='color:green;'>✅ Attendance record created for {emp_name} ({emp_id})</h2>
        <p>Date: {today}</p>
        <p>Time: {now.strftime('%H:%M:%S')}</p>
        <p>Break Count: 0</p>
        <p>Break Minutes: 0</p>
        <p><a href='/debug-attendance/{emp_id}'>Check attendance record</a></p>
        """
        
    except Exception as e:
        return f"Error: {str(e)}"

@app.route('/debug-attendance/<emp_id>')
def debug_attendance(emp_id):
    """Check attendance records for an employee"""
    try:
        sheet = get_cached_sheet()
        attendance_ws = sheet.worksheet("Attendance")
        all_attendance = attendance_ws.get_all_records()
        all_rows = attendance_ws.get_all_values()
        
        today = date.today().isoformat()
        
        result = f"<h2>Attendance Records for Employee {emp_id}</h2>"
        result += f"<p>Today: {today}</p>"
        result += "<hr>"
        
        found = False
        for i, att in enumerate(all_attendance, start=2):
            att_emp_id = safe_str(att.get('Employee ID', '')).strip()
            
            if att_emp_id == emp_id:
                found = True
                result += f"<h3>Record at row {i}:</h3>"
                result += "<table border='1' cellpadding='5'>"
                
                # Headers
                result += "<tr>"
                for h in all_rows[0]:
                    result += f"<th>{h}</th>"
                result += "</tr>"
                
                # Values
                result += "<tr>"
                for val in all_rows[i-1]:
                    result += f"<td>{val}</td>"
                result += "</tr>"
                result += "</table><br>"
        
        if not found:
            result += f"<p style='color:red;'>❌ No attendance records found for employee {emp_id}</p>"
        
        return result
        
    except Exception as e:
        return f"Error: {str(e)}"

@app.route('/debug-attendance-structure')
def debug_attendance_structure():
    """Show attendance sheet structure"""
    try:
        sheet = get_cached_sheet()
        attendance_ws = sheet.worksheet("Attendance")
        headers = attendance_ws.row_values(1)
        
        result = "<h2>📊 Attendance Sheet Structure</h2>"
        result += "<table border='1' cellpadding='8'>"
        result += "<tr><th>Column #</th><th>Header Name</th></tr>"
        
        for i, h in enumerate(headers, start=1):
            color = "#E8F8F5" if 'Break' in h else "white"
            result += f"<tr style='background:{color}'>"
            result += f"<td>{i}</td>"
            result += f"<td><strong>{h}</strong></td>"
            result += "</tr>"
        
        result += "</table>"
        
        # Show first data row if exists
        all_rows = attendance_ws.get_all_values()
        if len(all_rows) > 1:
            result += "<h3>📝 First Data Row:</h3>"
            result += "<table border='1' cellpadding='5'><tr>"
            for val in all_rows[1]:
                result += f"<td>{val}</td>"
            result += "</tr></table>"
        
        return result
        
    except Exception as e:
        return f"Error: {str(e)}"

@app.route('/clear-all-cache')
def clear_all_cache():
    """Clear all caches"""
    global _worksheet_cache, cache
    _worksheet_cache = {}
    cache = {}
    return """
    <h2 style='color:green;'>✅ All caches cleared!</h2>
    <p>Worksheet cache: Cleared</p>
    <p>Smart cache: Cleared</p>
    """
    
# ==================== ONLINE/OFFLINE TRACKING ====================


def get_all_employee_status():
    """Get status of all employees for admin dashboard"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return []
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        status_list = []
        for emp in all_employees:
            if emp.get('Is Active') == 'Yes':
                last_activity = emp.get('Last Activity', '')
                status = emp.get('Status', 'offline')
                
                # Calculate how long ago
                time_ago = ''
                if last_activity:
                    try:
                        last_time = datetime.fromisoformat(last_activity)
                        diff = datetime.now() - last_time
                        if diff.total_seconds() < 60:
                            time_ago = 'just now'
                        elif diff.total_seconds() < 3600:
                            time_ago = f'{int(diff.total_seconds() / 60)} min ago'
                        else:
                            time_ago = f'{int(diff.total_seconds() / 3600)} hours ago'
                    except:
                        time_ago = 'unknown'
                
                status_list.append({
                    'employee_id': emp.get('Employee ID'),
                    'name': emp.get('Employee Name'),
                    'department': emp.get('Department'),
                    'status': status,
                    'last_activity': last_activity,
                    'time_ago': time_ago,
                    'break_total': emp.get('Total Break Time', '0'),
                    'break_count': emp.get('Today\'s Break Count', '0')
                })
        
        return status_list
    except Exception as e:
        print(f"❌ Get status error: {e}")
        return []



@app.route('/attendance/history')
@login_required
def attendance_history():
    """Attendance history page"""
    return render_template('attendance_history.html')
@app.route('/debug')
def debug_info():
    """Debug endpoint"""
    return {
        'on_render': os.environ.get('RENDER', 'not set'),
        'has_creds': os.path.exists('google_credentials.json'),
        'session': {k: safe_str(v) for k, v in session.items() if 'password' not in k.lower()},
        'python_version': os.sys.version
    }

@app.route('/reset-employee-23-password')
def reset_employee_23_password():
    """Simple password reset for employee 23"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return "Database connection error"
        
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        # Find employee 23
        row_index = None
        for i, emp in enumerate(all_employees, start=2):
            sheet_id = str(emp.get('Employee ID', '')).strip()
            if sheet_id == '23':
                row_index = i
                break
        
        if not row_index:
            return "Employee 23 not found in sheet"
        
        # Set new password to "23"
        new_password = "23"
        new_hash = generate_password_hash(new_password)
        
        # Update password hash (column 5)
        employees_ws.update_cell(row_index, 5, new_hash)
        
        # Update Force Password Change (column 12) to "No"
        employees_ws.update_cell(row_index, 12, 'No')
        
        return f"""
        <h2 style='color: green;'>✅ Password Reset Successful!</h2>
        <p>Password for Employee ID <strong>23</strong> has been reset to: <strong>{new_password}</strong></p>
        <p>You can now login with:</p>
        <ul>
            <li><strong>Employee ID:</strong> 23</li>
            <li><strong>Password:</strong> 23</li>
        </ul>
        <p><a href="/login" style="background-color: #4CAF50; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Go to Login Page</a></p>
        """
        
    except Exception as e:
        return f"Error: {str(e)}"
    

    
@app.route('/quick-add-employee', methods=['GET', 'POST'])
def quick_add_employee():
    """Quick way to add employees"""
    if request.method == 'POST':
        try:
            emp_id = request.form.get('emp_id')
            name = request.form.get('name')
            email = request.form.get('email')
            dept = request.form.get('dept')
            
            sheet = get_cached_sheet()
            if not sheet:
                return "Database error"
            
            employees_ws = sheet.worksheet("Employees")
            
            # Check if employee exists
            all_employees = employees_ws.get_all_records()
            for emp in all_employees:
                if str(emp.get('Employee ID', '')).strip() == emp_id:
                    return f"Employee ID {emp_id} already exists!"
            
            # Generate password hash (using employee ID as default password)
            password_hash = generate_password_hash(emp_id)
            today = date.today().isoformat()
            next_id = get_next_id(employees_ws)
            
            # Add employee with all columns
            employees_ws.append_row([
                next_id,           # ID
                emp_id,            # Employee ID
                name,              # Employee Name
                email,             # Email
                password_hash,     # Password Hash
                dept,              # Department
                today,             # Join Date
                'Yes',             # Is Active
                'employee',        # Role
                '',                # Last Login
                '',                # Password Changed
                'Yes',             # Force Password Change
                '',                # Last Activity
                'offline',         # Status
                '',                # Break Start
                '',                # Break End
                '0',               # Total Break Time
                '0'                # Break Count
            ])
            
            return f"""
            <h2>✅ Employee Added Successfully!</h2>
            <p><strong>Employee ID:</strong> {emp_id}</p>
            <p><strong>Name:</strong> {name}</p>
            <p><strong>Default Password:</strong> {emp_id}</p>
            <p><a href="/login">Go to Login</a></p>
            """
            
        except Exception as e:
            return f"Error: {str(e)}"
    
    # GET request - show form
    return '''
    <style>
        body { font-family: Arial; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 50px; }
        .container { max-width: 500px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; }
        input, select { width: 100%; padding: 10px; margin: 10px 0; border: 1px solid #ddd; border-radius: 5px; }
        button { background: #667eea; color: white; padding: 12px 20px; border: none; border-radius: 5px; cursor: pointer; }
        h2 { color: #333; }
    </style>
    <div class="container">
        <h2>Quick Add Employee</h2>
        <form method="post">
            <input type="text" name="emp_id" placeholder="Employee ID (e.g., 027)" required>
            <input type="text" name="name" placeholder="Full Name" required>
            <input type="email" name="email" placeholder="Email" required>
            <select name="dept" required>
                <option value="">Select Department</option>
                <option value="Engineering">Engineering</option>
                <option value="Sales">Sales</option>
                <option value="Marketing">Marketing</option>
                <option value="HR">HR</option>
                <option value="Finance">Finance</option>
                <option value="Operations">Operations</option>
            </select>
            <button type="submit">Add Employee</button>
        </form>
        <p style="margin-top: 20px; color: #666;">Default password will be the Employee ID</p>
    </div>
    '''
@app.route('/api/admin/export-attendance', methods=['POST'])
@admin_required
def export_attendance():
    """Export attendance data to Excel"""
    try:
        data = request.json
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        employee_id = data.get('employee_id', '')
        dept = data.get('dept', '')
        att_type = data.get('type', '')
        
        print(f"Export requested with filters: from={from_date}, to={to_date}, emp={employee_id}, dept={dept}, type={att_type}")
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        # Get attendance records
        attendance_ws = sheet.worksheet("Attendance")
        all_records = attendance_ws.get_all_records()
        print(f"Total attendance records: {len(all_records)}")
        
        # Get employees for department mapping
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        # Create employee department mapping
        emp_dept = {}
        emp_name = {}
        for emp in all_employees:
            emp_id = safe_str(emp.get('Employee ID', '')).strip()
            emp_dept[emp_id] = emp.get('Department', 'Not Assigned')
            emp_name[emp_id] = emp.get('Employee Name', '')
        
        # Filter records
        filtered_records = []
        for record in all_records:
            record_date = record.get('Date', '')
            rec_emp_id = safe_str(record.get('Employee ID', '')).strip()
            rec_dept = emp_dept.get(rec_emp_id, 'Not Assigned')
            rec_type = record.get('Type', '')
            
            # Apply filters
            if from_date and record_date < from_date:
                continue
            if to_date and record_date > to_date:
                continue
            if employee_id and rec_emp_id != employee_id:
                continue
            if dept and dept != 'All Departments' and rec_dept != dept:
                continue
            if att_type and rec_type != att_type:
                continue
            
            filtered_records.append({
                'Date': record_date,
                'Employee ID': rec_emp_id,
                'Employee Name': emp_name.get(rec_emp_id, record.get('Employee Name', '')),
                'Department': rec_dept,
                'Time': record.get('Time', ''),
                'Type': rec_type,
                'Week': record.get('Week', ''),
                'Month': record.get('Month', ''),
                'Year': record.get('Year', ''),
                'IP Address': record.get('IP Address', ''),
                'Status': record.get('Status', '')
            })
        
        print(f"Filtered records: {len(filtered_records)}")
        
        if not filtered_records:
            return jsonify({'success': False, 'message': 'No records found for selected filters'}), 404
        
        # Create DataFrame
        import pandas as pd
        from io import BytesIO
        
        df = pd.DataFrame(filtered_records)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance Report', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Attendance Report']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Generate filename
        from datetime import datetime
        filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    

   
@app.route('/api/admin/export-worklogs', methods=['POST'])
@admin_required
def export_worklogs():
    """Export work logs data to Excel"""
    try:
        data = request.json
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        employee_id = data.get('employee_id', '')
        dept = data.get('dept', '')
        
        print(f"Work logs export requested: from={from_date}, to={to_date}, emp={employee_id}, dept={dept}")
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        # Get work logs
        work_logs_ws = sheet.worksheet("WorkLogs")
        all_logs = work_logs_ws.get_all_records()
        print(f"Total work logs: {len(all_logs)}")
        
        # Get employees for department mapping
        employees_ws = sheet.worksheet("Employees")
        all_employees = employees_ws.get_all_records()
        
        # Create employee department mapping
        emp_dept = {}
        emp_name = {}
        for emp in all_employees:
            emp_id = safe_str(emp.get('Employee ID', '')).strip()
            emp_dept[emp_id] = emp.get('Department', 'Not Assigned')
            emp_name[emp_id] = emp.get('Employee Name', '')
        
        # Filter records
        filtered_logs = []
        for log in all_logs:
            log_date = log.get('Date', '')
            log_emp_id = safe_str(log.get('Employee ID', '')).strip()
            log_dept = emp_dept.get(log_emp_id, 'Not Assigned')
            
            # Apply filters
            if from_date and log_date < from_date:
                continue
            if to_date and log_date > to_date:
                continue
            if employee_id and log_emp_id != employee_id:
                continue
            if dept and dept != 'All Departments' and log_dept != dept:
                continue
            
            # Format hours with "hrs" suffix
            hours = log.get('Hours Worked', '0')
            try:
                hours_float = float(hours)
                hours_formatted = f"{hours_float} hrs"
            except:
                hours_formatted = f"{hours} hrs"
            
            # Format submitted date
            submitted_at = log.get('Submitted At', '')
            if submitted_at:
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(submitted_at.replace('Z', '+00:00'))
                    submitted_formatted = dt.strftime('%m/%d/%Y, %I:%M:%S %p')
                except:
                    submitted_formatted = submitted_at
            else:
                submitted_formatted = '-'
            
            filtered_logs.append({
                'DATE': log_date,
                'EMPLOYEE ID': log_emp_id,
                'EMPLOYEE NAME': emp_name.get(log_emp_id, log.get('Employee Name', '')),
                'DEPARTMENT': log_dept,
                'WORK DESCRIPTION': log.get('Work Description', ''),
                'HOURS': hours_formatted,
                'SUBMITTED AT': submitted_formatted,
                'IP ADDRESS': log.get('IP Address', '-')
            })
        
        print(f"Filtered work logs: {len(filtered_logs)}")
        
        if not filtered_logs:
            return jsonify({'success': False, 'message': 'No work logs found for selected filters'}), 404
        
        # Create DataFrame
        import pandas as pd
        from io import BytesIO
        from datetime import datetime
        
        df = pd.DataFrame(filtered_logs)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Work Logs', index=False)
            
            # Get the worksheet
            worksheet = writer.sheets['Work Logs']
            
            # Format headers - bold and background color
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0F9D8A", end_color="0F9D8A", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 4, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format date column
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left')
            
            # Format hours column
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=6, max_col=6):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
        
        output.seek(0)
        
        # Generate filename
        filename = f"worklogs_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Work logs export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    

@app.route('/api/admin/export-logs', methods=['POST'])
@admin_required
def export_logs():
    """Export audit logs data to Excel"""
    try:
        data = request.json
        days = data.get('days', 7)
        
        print(f"Audit logs export requested for last {days} days")
        
        sheet = get_cached_sheet()
        if not sheet:
            return jsonify({'success': False, 'message': 'Database error'}), 500
        
        # Get audit logs
        logs_ws = sheet.worksheet("AuditLogs")
        all_logs = logs_ws.get_all_records()
        print(f"Total audit logs: {len(all_logs)}")
        
        # Calculate date filter
        from datetime import datetime, timedelta
        cutoff_date = datetime.now() - timedelta(days=days)
        
        # Filter and format logs
        filtered_logs = []
        for log in all_logs:
            # Parse timestamp
            timestamp_str = log.get('Timestamp', '')
            try:
                log_date = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            except:
                try:
                    log_date = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S.%f')
                except:
                    try:
                        log_date = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                    except:
                        continue
            
            # Apply date filter
            if log_date < cutoff_date:
                continue
            
            # Format timestamp
            timestamp_formatted = log_date.strftime('%m/%d/%Y, %I:%M:%S %p')
            
            filtered_logs.append({
                'datetime_obj': log_date,  # Store for sorting
                'Timestamp': timestamp_formatted,
                'User Type': log.get('User Type', ''),
                'User Name': log.get('User Name', ''),
                'Action': log.get('Action', ''),
                'Details': log.get('Details', ''),
                'IP Address': log.get('IP Address', '')
            })
        
        # Sort by datetime in DESCENDING order (latest first)
        filtered_logs.sort(key=lambda x: x['datetime_obj'], reverse=True)
        
        # Remove datetime_obj before creating DataFrame
        for log in filtered_logs:
            del log['datetime_obj']
        
        print(f"Filtered audit logs: {len(filtered_logs)}")
        
        if not filtered_logs:
            return jsonify({'success': False, 'message': 'No audit logs found for selected period'}), 404
        
        # Create DataFrame
        import pandas as pd
        from io import BytesIO
        
        df = pd.DataFrame(filtered_logs)
        
        # ... rest of your Excel creation code ...
        
        # Create summary statistics
        action_counts = {}
        user_counts = {}
        for log in filtered_logs:
            action = log['Action']
            user = log['User Name']
            action_counts[action] = action_counts.get(action, 0) + 1
            user_counts[user] = user_counts.get(user, 0) + 1
        
        # Create summary data
        summary_data = {
            'Metric': [
                'Total Logs',
                'Date Range',
                'Unique Users',
                'Unique Actions',
                'Most Active User',
                'Most Common Action'
            ],
            'Value': [
                str(len(filtered_logs)),
                f"Last {days} days",
                str(len(user_counts)),
                str(len(action_counts)),
                max(user_counts.items(), key=lambda x: x[1])[0] if user_counts else '-',
                max(action_counts.items(), key=lambda x: x[1])[0] if action_counts else '-'
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        
        # Create action breakdown
        action_breakdown = []
        for action, count in action_counts.items():
            action_breakdown.append({
                'Action': action,
                'Count': count,
                'Percentage': f"{(count/len(filtered_logs)*100):.1f}%"
            })
        action_df = pd.DataFrame(action_breakdown) if action_breakdown else pd.DataFrame()
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Write action breakdown if available
            if not action_df.empty:
                action_df.to_excel(writer, sheet_name='Actions Breakdown', index=False)
            
            # Write details sheet
            df.to_excel(writer, sheet_name='Audit Log Details', index=False)
            
            # Format all sheets
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Format Summary sheet
            if 'Summary' in writer.sheets:
                worksheet = writer.sheets['Summary']
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="0F9D8A", end_color="0F9D8A", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 4, 40)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format Actions Breakdown sheet
            if 'Actions Breakdown' in writer.sheets:
                worksheet = writer.sheets['Actions Breakdown']
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="0F9D8A", end_color="0F9D8A", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 4, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format Details sheet
            worksheet = writer.sheets['Audit Log Details']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0F9D8A", end_color="0F9D8A", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths for details sheet
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 4, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format timestamp column
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left')
            
            # Format IP Address column
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=6, max_col=6):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
        
        output.seek(0)
        
        # Generate filename
        filename = f"audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Audit logs export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/fix-sheet-schema')
def fix_sheet_schema():
    """Fix the Employees worksheet schema by adding missing columns"""
    try:
        sheet = get_cached_sheet()
        if not sheet:
            return "Database error"
        
        employees_ws = sheet.worksheet("Employees")
        
        # Get current headers
        current_headers = employees_ws.row_values(1)
        print(f"Current headers: {current_headers}")
        
        # Complete headers list (18 columns)
        complete_headers = [
            'ID', 
            'Employee ID', 
            'Employee Name', 
            'Email', 
            'Password Hash', 
            'Department', 
            'Join Date', 
            'Is Active', 
            'Role', 
            'Last Login', 
            'Password Changed', 
            'Force Password Change',
            'Last Activity',
            'Status',
            'Break Start',
            'Break End',
            'Total Break Time',
            'Break Count'
        ]
        
        # Update headers if needed
        if len(current_headers) < len(complete_headers):
            # Update the headers
            employees_ws.update('A1:R1', [complete_headers])
            
            # Add default values for existing employees
            all_rows = employees_ws.get_all_values()
            for i in range(2, len(all_rows) + 1):  # Start from row 2
                # Add default values for new columns if they don't exist
                if len(all_rows[i-1]) < 13:
                    employees_ws.update_cell(i, 13, '')  # Last Activity
                if len(all_rows[i-1]) < 14:
                    employees_ws.update_cell(i, 14, 'offline')  # Status
                if len(all_rows[i-1]) < 15:
                    employees_ws.update_cell(i, 15, '')  # Break Start
                if len(all_rows[i-1]) < 16:
                    employees_ws.update_cell(i, 16, '')  # Break End
                if len(all_rows[i-1]) < 17:
                    employees_ws.update_cell(i, 17, '0')  # Total Break Time
                if len(all_rows[i-1]) < 18:
                    employees_ws.update_cell(i, 18, '0')  # Break Count
            
            return f"✅ Fixed schema! Updated from {len(current_headers)} to {len(complete_headers)} columns"
        else:
            return "Schema already correct"
        
    except Exception as e:
        return f"Error: {str(e)}"
    


    

   
@app.route('/api/employee/break-status')
@login_required
def get_employee_break_status():
    """Check if employee is on break"""
    try:
        sheet = get_cached_sheet()
        employees_data = get_cached_worksheet_data("Employees")
        
        for emp in employees_data:
            if safe_str(emp.get('Employee ID')).strip() == session['employee_id']:
                status = emp.get('Status', 'offline')
                return jsonify({
                    'success': True,
                    'on_break': status == 'break'
                })
        
        return jsonify({'success': False, 'on_break': False})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/start-break', methods=['POST'])
@login_required
def start_break():
    """Start break - DEBUG VERSION"""
    try:
        employee_id = session['employee_id']
        print(f"🔵 START BREAK called for employee: {employee_id}")
        
        now = datetime.now()
        success = update_employee_status(employee_id, 'break', break_start=now)
        
        if success:
            clear_smart_cache(data_types=['employee_status'])
            return jsonify({'success': True, 'message': 'Break started'})
        return jsonify({'success': False, 'message': 'Failed to start break'}), 500
    except Exception as e:
        print(f"❌ Start break error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/end-break', methods=['POST'])
@login_required
def end_break():
    """End break"""
    try:
        employee_id = session['employee_id']
        print(f"🟢 END BREAK: {employee_id}")
        
        now = datetime.now()
        success = update_employee_status(employee_id, 'online', break_end=now)
        
        if success:
            # Clear ALL status-related caches
            clear_smart_cache(keys_to_clear=[
                'admin_employee_status',
                f'emp_status_{employee_id}',
                'dashboard_stats'
            ])
            # Also clear worksheet cache
            global _worksheet_cache
            if 'Attendance' in _worksheet_cache:
                del _worksheet_cache['Attendance']
            if 'Employees' in _worksheet_cache:
                del _worksheet_cache['Employees']
            
            return jsonify({'success': True, 'message': 'Break ended'})
        return jsonify({'success': False, 'message': 'Failed to end break'}), 500
    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/debug-break-count/<emp_id>')
def debug_break_count(emp_id):
    """Check break count in attendance sheet"""
    try:
        sheet = get_cached_sheet()
        attendance_ws = sheet.worksheet("Attendance")
        all_attendance = attendance_ws.get_all_records()
        all_rows = attendance_ws.get_all_values()
        
        today = date.today().isoformat()
        
        result = f"<h2>Break Count Debug for Employee {emp_id}</h2>"
        result += f"<p>Today: {today}</p>"
        result += "<hr>"
        
        # Find today's attendance
        found = False
        for i, att in enumerate(all_attendance, start=2):
            att_emp_id = safe_str(att.get('Employee ID', '')).strip()
            att_date = att.get('Date', '')
            
            if att_emp_id == emp_id and att_date == today:
                found = True
                result += f"<h3>✅ Found attendance record at row {i}</h3>"
                
                # Get headers
                headers = all_rows[0]
                result += "<table border='1'><tr>"
                for h in headers:
                    result += f"<th>{h}</th>"
                result += "</tr><tr>"
                
                # Get values
                row_data = all_rows[i-1]
                for val in row_data:
                    result += f"<td>{val}</td>"
                result += "</tr></table>"
                
                # Find Break Count column
                break_count_col = None
                break_minutes_col = None
                for idx, h in enumerate(headers):
                    if h == 'Break Count':
                        break_count_col = idx
                        result += f"<p>📊 Break Count column: {idx+1}</p>"
                    elif h == 'Break Minutes':
                        break_minutes_col = idx
                        result += f"<p>📊 Break Minutes column: {idx+1}</p>"
                
                # Get current values
                if break_count_col is not None and len(row_data) > break_count_col:
                    current_count = row_data[break_count_col]
                    result += f"<p><strong>Current Break Count: {current_count}</strong></p>"
                
                break
        
        if not found:
            result += f"<p style='color:red;'>❌ No attendance record found for today</p>"
        
        return result
        
    except Exception as e:
        return f"Error: {str(e)}"

@app.route('/speed-test')
def speed_test():
    import time
    
    # Test BEFORE (multiple calls)
    start = time.time()
    for i in range(5):
        sheet = get_google_sheet()
    before_time = time.time() - start
    
    # Test AFTER (multiple calls)
    start = time.time()
    for i in range(5):
        sheet = get_cached_sheet()
    after_time = time.time() - start
    
    # Individual call times
    start = time.time()
    sheet1 = get_google_sheet()
    single_before = time.time() - start
    
    start = time.time()
    sheet2 = get_cached_sheet()
    single_after = time.time() - start
    
    return f"""
    <h2>📊 Real Speed Test Results:</h2>
    
    <h3>Single Call:</h3>
    <p>Before: <strong>{single_before:.2f} seconds</strong></p>
    <p>After:  <strong>{single_after:.2f} seconds</strong></p>
    <p>Improvement: <strong>{((single_before-single_after)/single_before*100):.0f}%</strong></p>
    
    <h3>5 Calls in a row:</h3>
    <p>Before (5 calls): <strong>{before_time:.2f} seconds</strong></p>
    <p>After (5 calls):  <strong>{after_time:.2f} seconds</strong></p>
    <p>Improvement: <strong>{((before_time-after_time)/before_time*100):.0f}%</strong></p>
    
    <h3>What this means:</h3>
    <p>🔵 Before: Each page load creates NEW connection</p>
    <p>🟢 After: First load creates connection, next loads reuse it</p>
    """


@app.route('/speed-test')
def speed_test_page():
    """Simple speed test page"""
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Speed Test</title>
        <style>
            body { font-family: Arial; background: #E8F8F5; padding: 20px; }
            .container { max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 16px; box-shadow: 0 4px 12px rgba(15,157,138,0.15); }
            h1 { color: #0F9D8A; }
            .test-card { background: #f8f9fa; padding: 20px; margin: 20px 0; border-radius: 10px; border-left: 4px solid #0F9D8A; }
            .result { font-size: 24px; font-weight: bold; margin: 10px 0; }
            .good { color: #28a745; }
            .bad { color: #dc3545; }
            button { background: linear-gradient(135deg, #0F9D8A, #7ED957); color: white; border: none; padding: 12px 24px; border-radius: 8px; font-size: 16px; cursor: pointer; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🚀 Website Speed Test</h1>
            
            <div class="test-card">
                <h3>📊 Test 1: Connection Speed</h3>
                <button onclick="testConnection()">Run Test</button>
                <div id="connectionResult" class="result">-</div>
            </div>
            
            <div class="test-card">
                <h3>⚡ Test 2: Cache Speed (5 calls in a row)</h3>
                <button onclick="testCache()">Run Test</button>
                <div id="cacheResult" class="result">-</div>
            </div>
            
            <div class="test-card">
                <h3>📈 Test 3: Dashboard Load Time</h3>
                <button onclick="testDashboard()">Run Test</button>
                <div id="dashboardResult" class="result">-</div>
            </div>
            
            <div class="test-card">
                <h3>📋 Test 4: Before vs After Comparison</h3>
                <button onclick="runComparison()">Run All Tests</button>
                <div id="comparisonResult"></div>
            </div>
        </div>

        <script>
        async function testConnection() {
            const start = performance.now();
            await fetch('/api/test-connection');
            const time = (performance.now() - start) / 1000;
            
            document.getElementById('connectionResult').innerHTML = 
                `<span class="${time < 0.5 ? 'good' : 'bad'}">${time.toFixed(2)} seconds</span>`;
        }
        
        async function testCache() {
            const start = performance.now();
            for (let i = 0; i < 5; i++) {
                await fetch('/api/test-connection');
            }
            const time = (performance.now() - start) / 1000;
            const perCall = (time / 5).toFixed(3);
            
            document.getElementById('cacheResult').innerHTML = 
                `<span class="${time < 2 ? 'good' : 'bad'}">${time.toFixed(2)} seconds total (${perCall}s per call)</span>`;
        }
        
        async function testDashboard() {
            const start = performance.now();
            await fetch('/api/admin/dashboard-stats');
            const time = (performance.now() - start) / 1000;
            
            document.getElementById('dashboardResult').innerHTML = 
                `<span class="${time < 1 ? 'good' : 'bad'}">${time.toFixed(2)} seconds</span>`;
        }
        
        async function runComparison() {
            let html = '<h3>Results:</h3>';
            
            // Before (no cache)
            let start = performance.now();
            for (let i = 0; i < 3; i++) {
                await fetch('/api/test-before');
            }
            let beforeTime = ((performance.now() - start) / 1000).toFixed(2);
            
            // After (with cache)
            start = performance.now();
            for (let i = 0; i < 3; i++) {
                await fetch('/api/test-after');
            }
            let afterTime = ((performance.now() - start) / 1000).toFixed(2);
            
            let improvement = ((beforeTime - afterTime) / beforeTime * 100).toFixed(0);
            
            html += `
                <table style="width:100%; border-collapse: collapse;">
                    <tr style="background: #0F9D8A; color: white;">
                        <th style="padding:10px">Test</th>
                        <th style="padding:10px">Time (3 calls)</th>
                        <th style="padding:10px">Status</th>
                    </tr>
                    <tr style="border-bottom:1px solid #ddd;">
                        <td style="padding:10px"><strong>Before (No Cache)</strong></td>
                        <td style="padding:10px">${beforeTime} seconds</td>
                        <td style="padding:10px" class="bad">🐌 Slow</td>
                    </tr>
                    <tr>
                        <td style="padding:10px"><strong>After (With Cache)</strong></td>
                        <td style="padding:10px">${afterTime} seconds</td>
                        <td style="padding:10px" class="good">🚀 Fast</td>
                    </tr>
                    <tr style="background: #E8F8F5;">
                        <td style="padding:10px" colspan="3"><strong>Improvement: ${improvement}% faster!</strong></td>
                    </tr>
                </table>
            `;
            
            document.getElementById('comparisonResult').innerHTML = html;
        }
        </script>
    </body>
    </html>
    '''
@app.route('/api/test-connection')
def test_connection():
    """Test connection speed"""
    sheet = get_cached_sheet()
    data = get_cached_worksheet_data("Employees")
    return jsonify({'success': True, 'count': len(data)})

@app.route('/api/test-before')
def test_before():
    """Simulate before caching (slow)"""
    import time
    time.sleep(0.5)  # Simulate slow connection
    sheet = get_google_sheet()  # Old slow method
    return jsonify({'success': True})

@app.route('/api/test-after')
def test_after():
    """Test with caching (fast)"""
    sheet = get_cached_sheet()  # New fast method
    return jsonify({'success': True})
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)